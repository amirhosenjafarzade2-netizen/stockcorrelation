# excel_export_improved.py
# Enhanced Excel Export Module with Multi-Ticker, Advanced Analytics, and Professional Formatting
#
# v2 improvements over the original:
#   AESTHETICS  - hyperlinked Table of Contents / cover sheet, consistent Calibri styling,
#                 thin borders around every table, icon-set arrows on return figures,
#                 "back to contents" navigation link on every sheet, print-ready page setup
#                 (fit-to-width, repeating header row, landscape) on every sheet.
#   MORE INFO   - company profile snapshot (sector/industry/market cap/employees/website),
#                 52-week range & beta, a "Growth of $100" rebased comparison chart,
#                 moving averages (SMA20/50/200) and RSI(14) on the price sheet.
#   MORE ABILITY- auto-generated plain-English "Key Takeaways" bullets per ticker,
#                 correlation heatmap sheet for portfolio mode, RSI overbought/oversold
#                 conditional formatting, rebased multi-ticker performance chart.

import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Dict, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings
warnings.filterwarnings('ignore')

# Check for openpyxl (required for Excel operations)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, ScatterChart, Series, Reference
    from openpyxl.chart.trendline import Trendline
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, IconSetRule
    from openpyxl.worksheet.hyperlink import Hyperlink
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    st.warning("⚠️ openpyxl not available. Install with: pip install openpyxl")


class ExcelExporter:
    """Handles Excel export with multiple sheets, advanced formatting, charts, and tables"""

    def __init__(self):
        """Initialize the exporter with color schemes and formats"""
        self.colors = {
            'primary': 'FF4472C4',
            'secondary': 'FF70AD47',
            'accent': 'FFFFC000',
            'danger': 'FFE74C3C',
            'success': 'FF2ECC71',
            'light_gray': 'FFF2F2F2',
            'zebra': 'FFF7F9FC',
            'dark_gray': 'FF7F7F7F',
            'navy': 'FF1F3864',
            'gold_text': 'FFB8860B',
        }
        # Tab colors (hex, no leading 'FF' needed for sheet_properties.tabColor)
        self.tab_colors = {
            'contents': '1F3864',
            'summary': '4472C4',
            'price': '2E75B6',
            'analysis': '70AD47',
            'comparison': '4472C4',
            'metrics': 'FFC000',
            'dividends': '9DC3E6',
            'benchmark': 'C00000',
            'seasonality': '7030A0',
            'correlation': '375623',
        }
        self.base_font_name = 'Calibri'
        self._table_names_used = set()
        self._sheet_registry: List[Dict] = []

    # ------------------------------------------------------------------
    # Generic styling helpers
    # ------------------------------------------------------------------
    def _center_all_cells(self, ws):
        """Center-align (horizontal + vertical) every populated cell in the sheet.
        Used only for key/value style summary blocks, not wide data tables."""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def _style_header_row(self, ws, row_idx, n_cols, color_key='secondary', start_col=1):
        """Bold white-on-color header, centered, for a single row."""
        fill = PatternFill(start_color=self.colors[color_key], end_color=self.colors[color_key], fill_type='solid')
        thin = Side(style='thin', color='FFFFFFFF')
        for c in range(start_col, start_col + n_cols):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = Font(name=self.base_font_name, bold=True, color='FFFFFF')
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(bottom=Side(style='thin', color='FFFFFF'))

    def _apply_zebra(self, ws, first_row, last_row, first_col, last_col):
        """Alternate light-gray fill on data rows so wide tables are easier to scan."""
        fill = PatternFill(start_color=self.colors['zebra'], end_color=self.colors['zebra'], fill_type='solid')
        for r in range(first_row, last_row + 1):
            if (r - first_row) % 2 == 1:
                for c in range(first_col, last_col + 1):
                    ws.cell(row=r, column=c).fill = fill

    def _add_thin_border_box(self, ws, first_row, last_row, first_col, last_col):
        """Draw a light box border around a table so it reads as a discrete card, not
        just text floating on the sheet."""
        thin = Side(style='thin', color='FFD9D9D9')
        for r in range(first_row, last_row + 1):
            for c in range(first_col, last_col + 1):
                cell = ws.cell(row=r, column=c)
                top = thin if r == first_row else cell.border.top
                bottom = thin if r == last_row else cell.border.bottom
                left = thin if c == first_col else cell.border.left
                right = thin if c == last_col else cell.border.right
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    def _unique_table_name(self, base):
        """Excel table names must be unique per workbook and contain no spaces/symbols."""
        clean = ''.join(ch if ch.isalnum() else '_' for ch in base)
        name = clean
        i = 2
        while name in self._table_names_used:
            name = f"{clean}_{i}"
            i += 1
        self._table_names_used.add(name)
        return name

    def _add_excel_table(self, ws, ref, name_hint, style='TableStyleMedium9'):
        """Wrap a data range as a real Excel Table -> gives filter dropdowns + banded rows natively."""
        try:
            name = self._unique_table_name(name_hint)
            tbl = Table(displayName=name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name=style, showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tbl)
        except Exception:
            pass  # Non-fatal — sheet still works without the native Table wrapper

    def _autosize_columns(self, ws, n_cols, start_row=1, end_row=None, min_width=8, max_width=25, start_col=1):
        end_row = end_row or ws.max_row
        for idx in range(n_cols):
            col_idx = start_col + idx
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(start_row, end_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, min_width), max_width)

    def _format_numeric_column(self, ws, col_idx, first_row, last_row, fmt, align='right'):
        for r in range(first_row, last_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal=align, vertical='center')

    def _add_color_scale(self, ws, cell_range, colors=('F8696B', 'FFEB84', '63BE7B')):
        """Red-yellow-green 3-color scale, e.g. for returns or R^2 columns."""
        rule = ColorScaleRule(
            start_type='min', start_color=colors[0],
            mid_type='percentile', mid_value=50, mid_color=colors[1],
            end_type='max', end_color=colors[2]
        )
        ws.conditional_formatting.add(cell_range, rule)

    def _add_icon_set(self, ws, cell_range):
        """3-arrow icon set (up/flat/down) — a quicker visual read than color alone,
        especially useful for return columns scanned at a glance."""
        try:
            rule = IconSetRule(icon_style='3Arrows', type='percent', values=[0, 33, 67], showValue=True, reverse=False)
            ws.conditional_formatting.add(cell_range, rule)
        except Exception:
            pass

    def _set_tab_color(self, ws, key):
        try:
            ws.sheet_properties.tabColor = self.tab_colors.get(key, '000000')
        except Exception:
            pass

    def _apply_print_setup(self, ws, n_cols, orientation='landscape', repeat_row=3):
        """Make every sheet sane to print/PDF: fit-to-width, repeated header row,
        reasonable margins, and a footer with page numbers."""
        try:
            ws.page_setup.orientation = orientation
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_options.gridLines = False
            ws.page_margins.left = 0.4
            ws.page_margins.right = 0.4
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.oddFooter.center.text = "Page &P of &N"
            ws.oddFooter.right.text = "&D"
            if repeat_row:
                ws.print_title_rows = f"{repeat_row}:{repeat_row}"
            ws.sheet_view.showGridLines = False
        except Exception:
            pass

    def _register_sheet(self, sheet_name, title, tab_key, description=""):
        """Track every sheet we build so the Contents page can list & hyperlink to it,
        in the exact order sheets were created."""
        self._sheet_registry.append({
            'sheet_name': sheet_name, 'title': title,
            'tab_key': tab_key, 'description': description
        })

    def _add_nav_link(self, ws, n_cols, row=1):
        """Small '⌂ Contents' hyperlink tucked in the top-right corner of every sheet
        so users can jump back to the index without hunting for the tab."""
        try:
            col = max(n_cols + 2, 2)
            cell = ws.cell(row=row, column=col, value="⌂ Contents")
            cell.hyperlink = "#'Contents'!A1"
            cell.font = Font(name=self.base_font_name, size=9, color='FFFFFF', underline='single')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Workbook assembly
    # ------------------------------------------------------------------
    def create_workbook(self,
                       ticker_data: Dict[str, pd.DataFrame],
                       analysis_data: Optional[Dict[str, pd.DataFrame]] = None,
                       metadata: Optional[Dict] = None,
                       include_charts: bool = True,
                       metrics_data: Optional[Dict[str, pd.DataFrame]] = None,
                       dividends_data: Optional[Dict[str, Tuple[pd.Series, pd.Series]]] = None,
                       benchmark_data: Optional[Dict[str, Dict]] = None,
                       seasonality_data: Optional[Dict[str, Dict[str, pd.DataFrame]]] = None,
                       profile_data: Optional[Dict[str, Dict]] = None) -> BytesIO:
        """
        Create formatted Excel workbook with multiple sheets.

        Args:
            ticker_data: {ticker: price_df}
            analysis_data: {sheet_name: analysis_df}
            metadata: Export metadata (tickers, dates, etc.)
            include_charts: Whether to include charts
            metrics_data: {ticker: metrics_df} earnings vs fundamentals table per ticker
            dividends_data: {ticker: (dividends_series, splits_series)}
            benchmark_data: {ticker: {'benchmark_ticker': str, 'rel_df': df, 'stats': dict}}
            seasonality_data: {ticker: {'monthly': df, 'weekday': df}}
            profile_data: {ticker: {field: value}} company profile snapshot

        Returns:
            BytesIO buffer with Excel file
        """
        output = BytesIO()
        section_errors = []
        self._sheet_registry = []
        self._table_names_used = set()

        # Each sheet is built inside its own try/except so a failure in one sheet
        # (e.g. a chart error) never discards every other sheet.
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Summary Sheet
            if metadata:
                try:
                    self._create_summary_sheet(writer, metadata, ticker_data, profile_data)
                except Exception as e:
                    section_errors.append(f"Summary sheet: {e}")

            # 2. Price Data Sheets (one per ticker)
            for ticker, df in ticker_data.items():
                try:
                    self._create_price_sheet(writer, ticker, df, include_charts)
                except Exception as e:
                    section_errors.append(f"{ticker} price sheet: {e}")

            # 3. Analysis Sheets
            if analysis_data:
                for sheet_name, df in analysis_data.items():
                    try:
                        self._create_analysis_sheet(writer, sheet_name, df)
                    except Exception as e:
                        section_errors.append(f"{sheet_name} analysis sheet: {e}")

            # 4. Comparison Sheet (if multiple tickers)
            if len(ticker_data) > 1:
                try:
                    self._create_comparison_sheet(writer, ticker_data, metadata, include_charts)
                except Exception as e:
                    section_errors.append(f"Comparison sheet: {e}")

            # 4b. Correlation heatmap (if multiple tickers)
            if len(ticker_data) > 1:
                try:
                    self._create_correlation_sheet(writer, ticker_data)
                except Exception as e:
                    section_errors.append(f"Correlation sheet: {e}")

            # 5. Earnings vs Metrics Correlation Sheets
            if metrics_data:
                for ticker, mdf in metrics_data.items():
                    if mdf is not None and not mdf.empty:
                        try:
                            self._create_metrics_correlation_sheet(writer, ticker, mdf)
                        except Exception as e:
                            section_errors.append(f"{ticker} earnings/metrics sheet: {e}")

            # 6. Dividends & Splits Sheets
            if dividends_data:
                for ticker, (divs, splits) in dividends_data.items():
                    if (divs is not None and not divs.empty) or (splits is not None and not splits.empty):
                        try:
                            self._create_dividends_sheet(writer, ticker, divs, splits)
                        except Exception as e:
                            section_errors.append(f"{ticker} dividends sheet: {e}")

            # 7. Benchmark Comparison Sheets
            if benchmark_data:
                for ticker, bdata in benchmark_data.items():
                    if bdata and bdata.get('rel_df') is not None:
                        try:
                            self._create_benchmark_sheet(writer, ticker, bdata, include_charts)
                        except Exception as e:
                            section_errors.append(f"{ticker} benchmark sheet: {e}")

            # 8. Seasonality Sheets
            if seasonality_data:
                for ticker, sdata in seasonality_data.items():
                    if sdata:
                        try:
                            self._create_seasonality_sheet(writer, ticker, sdata, include_charts)
                        except Exception as e:
                            section_errors.append(f"{ticker} seasonality sheet: {e}")

            # 9. Table of Contents / cover — built last since it needs the full registry,
            #    then moved to the front of the tab order.
            try:
                self._create_toc_sheet(writer, metadata, ticker_data)
            except Exception as e:
                section_errors.append(f"Contents sheet: {e}")

            try:
                wb = writer.book
                if 'Contents' in wb.sheetnames:
                    toc_ws = wb['Contents']
                    wb._sheets.remove(toc_ws)
                    wb._sheets.insert(0, toc_ws)
                    wb.active = 0
            except Exception:
                pass

        if section_errors:
            for err in section_errors:
                st.warning(f"⚠️ Issue while building workbook — {err}")

        output.seek(0)
        return output

    def _create_basic_workbook(self, ticker_data, analysis_data, metadata):
        """Fallback basic Excel creation without advanced formatting"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if metadata:
                summary_df = pd.DataFrame({
                    'Parameter': ['Export Date', 'Tickers', 'Start Date', 'End Date', 'Frequency', 'Total Records'],
                    'Value': [
                        metadata.get('export_date', ''),
                        ', '.join(metadata.get('tickers', [])),
                        str(metadata.get('start_date', '')),
                        str(metadata.get('end_date', '')),
                        metadata.get('frequency', ''),
                        metadata.get('total_records', 0)
                    ]
                })
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

            for ticker, df in ticker_data.items():
                sheet_name = f"{ticker}_Prices"[:31]
                df.to_excel(writer, sheet_name=sheet_name)

            if analysis_data:
                for sheet_name, df in analysis_data.items():
                    df.to_excel(writer, sheet_name=sheet_name[:31])

        output.seek(0)
        return output

    # ------------------------------------------------------------------
    # Table of Contents / cover sheet
    # ------------------------------------------------------------------
    def _create_toc_sheet(self, writer, metadata, ticker_data):
        """A hyperlinked index/cover page — the first thing a user sees when they open
        the file. Gives the report a professional 'front matter' feel and makes the
        (potentially 15-20 sheet) workbook navigable instead of a wall of tabs."""
        wb = writer.book
        ws = wb.create_sheet('Contents')
        self._set_tab_color(ws, 'contents')
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:F3')
        ws['A1'] = "📊 Historical Stock Data Report"
        ws['A1'].font = Font(name=self.base_font_name, size=22, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['navy'], end_color=self.colors['navy'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        meta = metadata or {}
        subtitle = (f"{', '.join(meta.get('tickers', ticker_data.keys()))}  •  "
                    f"{meta.get('start_date', '')} → {meta.get('end_date', '')}  •  "
                    f"{meta.get('frequency', 'Daily')} data")
        ws.merge_cells('A4:F4')
        ws['A4'] = subtitle
        ws['A4'].font = Font(name=self.base_font_name, size=11, italic=True, color=self.colors['dark_gray'][2:])
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A5:F5')
        ws['A5'] = f"Generated {meta.get('export_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}"
        ws['A5'].font = Font(name=self.base_font_name, size=9, color='808080')
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')

        header_row = 7
        headers = ['#', 'Sheet', 'Contents', '']
        for i, h in enumerate(headers):
            c = ws.cell(row=header_row, column=i + 1, value=h)
        self._style_header_row(ws, header_row, 4, color_key='primary')

        row = header_row + 1
        for i, entry in enumerate(self._sheet_registry, 1):
            ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
            link_cell = ws.cell(row=row, column=2, value=entry['title'])
            link_cell.hyperlink = f"#'{entry['sheet_name']}'!A1"
            link_cell.font = Font(name=self.base_font_name, color='1155CC', underline='single', bold=True)
            ws.cell(row=row, column=3, value=entry.get('description', ''))
            swatch = ws.cell(row=row, column=4)
            swatch.fill = PatternFill(start_color=self.tab_colors.get(entry['tab_key'], '000000'),
                                       end_color=self.tab_colors.get(entry['tab_key'], '000000'), fill_type='solid')
            row += 1

        last_row = row - 1
        if last_row >= header_row + 1:
            self._apply_zebra(ws, header_row + 1, last_row, 1, 4)
            self._add_thin_border_box(ws, header_row, last_row, 1, 4)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 4
        for extra in 'EF':
            ws.column_dimensions[extra].width = 14

        footer_row = last_row + 3
        ws.merge_cells(f'A{footer_row}:F{footer_row}')
        ws[f'A{footer_row}'] = ("Data sourced via Yahoo Finance (yfinance). For informational purposes only — "
                                 "not investment advice.")
        ws[f'A{footer_row}'].font = Font(name=self.base_font_name, size=8, italic=True, color='999999')
        ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')

        ws.freeze_panes = 'A8'
        self._apply_print_setup(ws, 4, orientation='portrait', repeat_row=None)

    # ------------------------------------------------------------------
    # Individual sheets
    # ------------------------------------------------------------------
    def _create_summary_sheet(self, writer, metadata, ticker_data, profile_data=None):
        """Create formatted summary sheet"""
        summary_data = {
            'Parameter': [
                'Export Date & Time',
                'Number of Tickers',
                'Tickers',
                'Start Date',
                'End Date',
                'Period (Days)',
                'Data Frequency',
                'Total Data Points',
                'Average Points per Ticker'
            ],
            'Value': [
                metadata.get('export_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                len(metadata.get('tickers', [])),
                ', '.join(metadata.get('tickers', [])),
                str(metadata.get('start_date', '')),
                str(metadata.get('end_date', '')),
                (metadata.get('end_date') - metadata.get('start_date')).days if metadata.get('end_date') and metadata.get('start_date') else 'N/A',
                metadata.get('frequency', 'Daily'),
                metadata.get('total_records', 0),
                metadata.get('total_records', 0) // max(len(metadata.get('tickers', [])), 1)
            ]
        }

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=2)

        ws = writer.sheets['Summary']
        self._set_tab_color(ws, 'summary')
        ws.sheet_view.showGridLines = False

        ws['A1'] = 'Historical Stock Data Export'
        ws['A1'].font = Font(name=self.base_font_name, size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        ws.merge_cells('A1:D1')
        self._add_nav_link(ws, 3)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        self._style_header_row(ws, 3, 2)
        self._center_all_cells(ws)
        self._add_thin_border_box(ws, 3, 3 + len(summary_data['Parameter']), 1, 2)
        ws.freeze_panes = 'A4'

        cursor = len(summary_data['Parameter']) + 6

        if ticker_data:
            ws[f'A{cursor}'] = '📈 Quick Performance Summary'
            ws[f'A{cursor}'].font = Font(name=self.base_font_name, size=12, bold=True, color=self.colors['navy'][2:])

            perf_data = []
            for ticker, df in ticker_data.items():
                if len(df) > 1 and 'Close' in df.columns:
                    start_price = df['Close'].iloc[0]
                    end_price = df['Close'].iloc[-1]
                    total_return = ((end_price - start_price) / start_price) * 100

                    perf_data.append({
                        'Ticker': ticker,
                        'Start Price': round(float(start_price), 2),
                        'End Price': round(float(end_price), 2),
                        'Return (%)': round(float(total_return), 2)
                    })

            if perf_data:
                df_perf = pd.DataFrame(perf_data)
                perf_start_row = cursor + 2
                df_perf.to_excel(writer, sheet_name='Summary', index=False, startrow=cursor + 1)
                self._style_header_row(ws, perf_start_row, len(df_perf.columns))
                first_data_row = perf_start_row + 1
                last_data_row = perf_start_row + len(df_perf)
                self._format_numeric_column(ws, 2, first_data_row, last_data_row, '$#,##0.00')
                self._format_numeric_column(ws, 3, first_data_row, last_data_row, '$#,##0.00')
                self._format_numeric_column(ws, 4, first_data_row, last_data_row, '+0.00%;-0.00%')
                # Return values are plain percentages (e.g. 12.5), convert to fraction for % format
                for r in range(first_data_row, last_data_row + 1):
                    cell = ws.cell(row=r, column=4)
                    if isinstance(cell.value, (int, float)):
                        cell.value = cell.value / 100
                self._apply_zebra(ws, first_data_row, last_data_row, 1, 4)
                self._add_color_scale(ws, f'D{first_data_row}:D{last_data_row}')
                self._add_icon_set(ws, f'D{first_data_row}:D{last_data_row}')
                self._add_thin_border_box(ws, perf_start_row, last_data_row, 1, 4)
                cursor = last_data_row + 3
            else:
                cursor += 2

        # ---- Company profile snapshot (sector / industry / market cap / etc.) ----
        if profile_data:
            ws[f'A{cursor}'] = '🏢 Company Snapshot'
            ws[f'A{cursor}'].font = Font(name=self.base_font_name, size=12, bold=True, color=self.colors['navy'][2:])
            prof_header_row = cursor + 1
            fields = ['Ticker', 'Company', 'Sector', 'Industry', 'Market Cap', 'Employees',
                      '52W Low', '52W High', 'Beta', 'Div. Yield']
            for i, f in enumerate(fields):
                ws.cell(row=prof_header_row, column=i + 1, value=f)
            self._style_header_row(ws, prof_header_row, len(fields))

            r = prof_header_row + 1
            for ticker, prof in profile_data.items():
                if not prof:
                    continue
                ws.cell(row=r, column=1, value=ticker)
                ws.cell(row=r, column=2, value=prof.get('Company Name') or 'N/A')
                ws.cell(row=r, column=3, value=prof.get('Sector') or 'N/A')
                ws.cell(row=r, column=4, value=prof.get('Industry') or 'N/A')
                mc = prof.get('Market Cap')
                mc_cell = ws.cell(row=r, column=5, value=mc if mc else None)
                if mc:
                    mc_cell.number_format = '$#,##0,,"M"' if mc < 1e9 else '$#,##0,,,"B"'
                emp = prof.get('Employees')
                ws.cell(row=r, column=6, value=int(emp) if emp else None).number_format = '#,##0'
                low52 = prof.get('52W Low')
                ws.cell(row=r, column=7, value=round(low52, 2) if low52 else None).number_format = '$#,##0.00'
                high52 = prof.get('52W High')
                ws.cell(row=r, column=8, value=round(high52, 2) if high52 else None).number_format = '$#,##0.00'
                beta = prof.get('Beta')
                ws.cell(row=r, column=9, value=round(beta, 2) if beta else None).number_format = '0.00'
                dy = prof.get('Dividend Yield')
                dy_cell = ws.cell(row=r, column=10, value=(dy / 100 if dy and dy > 1 else dy))
                dy_cell.number_format = '0.00%'
                r += 1
            last_prof_row = r - 1
            if last_prof_row >= prof_header_row + 1:
                self._apply_zebra(ws, prof_header_row + 1, last_prof_row, 1, len(fields))
                self._add_thin_border_box(ws, prof_header_row, last_prof_row, 1, len(fields))
                self._autosize_columns(ws, len(fields), start_row=prof_header_row, end_row=last_prof_row, min_width=10, max_width=28)
                cursor = last_prof_row + 3

        # ---- Auto-generated plain-English takeaways ----
        if ticker_data:
            ws[f'A{cursor}'] = '💡 Key Takeaways'
            ws[f'A{cursor}'].font = Font(name=self.base_font_name, size=12, bold=True, color=self.colors['navy'][2:])
            bullets = self._generate_key_takeaways(ticker_data)
            for i, bullet in enumerate(bullets):
                ws.cell(row=cursor + 1 + i, column=1, value=f"•  {bullet}")
                ws.merge_cells(start_row=cursor + 1 + i, start_column=1, end_row=cursor + 1 + i, end_column=6)
                ws.cell(row=cursor + 1 + i, column=1).font = Font(name=self.base_font_name, size=10)
                ws.cell(row=cursor + 1 + i, column=1).alignment = Alignment(horizontal='left')

        self._apply_print_setup(ws, 4, orientation='portrait', repeat_row=None)

    def _generate_key_takeaways(self, ticker_data: Dict[str, pd.DataFrame]) -> List[str]:
        """Turn the numbers into a few plain-English sentences — best/worst performer,
        volatility flag, and any ticker sitting near its period high/low. Purely
        descriptive statements about the fetched data, not investment advice."""
        bullets = []
        perf = {}
        for ticker, df in ticker_data.items():
            if len(df) > 1 and 'Close' in df.columns:
                ret = (df['Close'].iloc[-1] / df['Close'].iloc[0] - 1) * 100
                vol = df['Close'].pct_change().std() * np.sqrt(252) * 100
                perf[ticker] = (ret, vol)

        if len(perf) > 1:
            best = max(perf.items(), key=lambda x: x[1][0])
            worst = min(perf.items(), key=lambda x: x[1][0])
            bullets.append(f"{best[0]} was the top performer over the period ({best[1][0]:+.1f}%), "
                            f"while {worst[0]} lagged the group ({worst[1][0]:+.1f}%).")
        elif len(perf) == 1:
            t, (ret, vol) = next(iter(perf.items()))
            direction = "gained" if ret >= 0 else "declined"
            bullets.append(f"{t} {direction} {abs(ret):.1f}% over the selected period with "
                            f"annualized volatility of {vol:.1f}%.")

        for ticker, (ret, vol) in perf.items():
            if vol > 45:
                bullets.append(f"{ticker} shows elevated volatility (~{vol:.0f}% annualized) — "
                                f"expect wider day-to-day price swings than a typical large-cap stock.")

        for ticker, df in ticker_data.items():
            if 'Close' in df.columns and len(df) > 5:
                current = df['Close'].iloc[-1]
                high = df['Close'].max()
                low = df['Close'].min()
                if high > 0 and (current / high) >= 0.98:
                    bullets.append(f"{ticker} is trading within 2% of its period high (${high:,.2f}).")
                elif low > 0 and (current / low) <= 1.02:
                    bullets.append(f"{ticker} is trading within 2% of its period low (${low:,.2f}).")

        if not bullets:
            bullets.append("Not enough price history in the selected window to summarize trends.")
        return bullets[:8]

    def _create_price_sheet(self, writer, ticker, df, include_charts=True):
        """Create formatted price data sheet with daily-return column, moving averages,
        RSI, native table, color scale, and an overlay chart."""
        df_out = df.copy()
        price_col = 'Adj Close' if 'Adj Close' in df_out.columns else ('Close' if 'Close' in df_out.columns else None)
        if price_col:
            df_out['Daily Return'] = df_out[price_col].pct_change()
            df_out['SMA 20'] = df_out[price_col].rolling(20).mean()
            df_out['SMA 50'] = df_out[price_col].rolling(50).mean()
            # RSI (14) — classic Wilder smoothing
            delta = df_out[price_col].diff()
            gain = delta.clip(lower=0)
            loss = -delta.clip(upper=0)
            avg_gain = gain.ewm(alpha=1/14, min_periods=14, adjust=False).mean()
            avg_loss = loss.ewm(alpha=1/14, min_periods=14, adjust=False).mean()
            rs = avg_gain / avg_loss.replace(0, np.nan)
            df_out['RSI (14)'] = 100 - (100 / (1 + rs))

        df_out.to_excel(writer, sheet_name=f"{ticker}_Prices"[:31], startrow=2)
        sheet_name = f"{ticker}_Prices"[:31]
        ws = writer.sheets[sheet_name]
        self._set_tab_color(ws, 'price')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, f"{ticker} — Price Data", 'price',
                              f"Daily OHLCV, moving averages, RSI, and return column ({len(df_out)} rows)")

        all_cols = ['Date'] + list(df_out.columns)
        n_cols = len(all_cols)

        ws['A1'] = f'{ticker} Historical Price Data'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols - 2, 1))}1')
        self._add_nav_link(ws, n_cols)

        for idx, col in enumerate(all_cols):
            ws.cell(row=3, column=idx + 1, value=col)
        self._style_header_row(ws, 3, n_cols)

        first_data_row, last_data_row = 4, len(df_out) + 3

        # Real numeric formats (not pre-formatted strings) so Excel can sort/filter/chart natively
        col_index = {name: i + 1 for i, name in enumerate(all_cols)}
        for money_col in ['Open', 'High', 'Low', 'Close', 'Adj Close', 'SMA 20', 'SMA 50']:
            if money_col in col_index:
                self._format_numeric_column(ws, col_index[money_col], first_data_row, last_data_row, '$#,##0.00')
        if 'Volume' in col_index:
            self._format_numeric_column(ws, col_index['Volume'], first_data_row, last_data_row, '#,##0')
        if 'Daily Return' in col_index:
            self._format_numeric_column(ws, col_index['Daily Return'], first_data_row, last_data_row, '+0.00%;-0.00%')
            ret_col_letter = get_column_letter(col_index['Daily Return'])
            self._add_color_scale(ws, f'{ret_col_letter}{first_data_row}:{ret_col_letter}{last_data_row}')
        if 'RSI (14)' in col_index:
            self._format_numeric_column(ws, col_index['RSI (14)'], first_data_row, last_data_row, '0.0')
            rsi_letter = get_column_letter(col_index['RSI (14)'])
            # Red = overbought (>70), green = oversold (<30) — flagged relative to neutral 50
            self._add_color_scale(ws, f'{rsi_letter}{first_data_row}:{rsi_letter}{last_data_row}',
                                   colors=('63BE7B', 'FFEB84', 'F8696B'))

        # Date column formatting/width
        ws.column_dimensions['A'].width = 14
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

        self._autosize_columns(ws, n_cols - 1, start_row=3, end_row=min(last_data_row, 103), start_col=2, min_width=10, max_width=16)
        self._apply_zebra(ws, first_data_row, last_data_row, 1, n_cols)
        self._add_thin_border_box(ws, 3, last_data_row, 1, n_cols)

        # Native Excel Table for filter/sort dropdowns (only when row count fits comfortably)
        if len(df_out) <= 100000:
            table_ref = f"A3:{get_column_letter(n_cols)}{last_data_row}"
            self._add_excel_table(ws, table_ref, f"{ticker}_Prices")

        # Summary statistics below the table
        stats_last_row = last_data_row
        if 'Close' in df_out.columns:
            stats_row = last_data_row + 3
            ws[f'A{stats_row}'] = 'Summary Statistics'
            ws[f'A{stats_row}'].font = Font(name=self.base_font_name, bold=True, size=11)

            current_rsi = df_out['RSI (14)'].iloc[-1] if 'RSI (14)' in df_out.columns else np.nan
            rsi_note = ('Overbought (>70)' if pd.notna(current_rsi) and current_rsi > 70 else
                        'Oversold (<30)' if pd.notna(current_rsi) and current_rsi < 30 else
                        'Neutral' if pd.notna(current_rsi) else 'N/A')

            stats = {
                'Current Price': f"${df_out['Close'].iloc[-1]:.2f}",
                'Period High': f"${df_out['High'].max():.2f}" if 'High' in df_out.columns else 'N/A',
                'Period Low': f"${df_out['Low'].min():.2f}" if 'Low' in df_out.columns else 'N/A',
                'Average Volume': f"{int(df_out['Volume'].mean()):,}" if 'Volume' in df_out.columns else 'N/A',
                'Total Return': f"{((df_out['Close'].iloc[-1] / df_out['Close'].iloc[0] - 1) * 100):+.2f}%" if len(df_out) > 1 else 'N/A',
                'Current RSI (14)': f"{current_rsi:.1f} — {rsi_note}" if pd.notna(current_rsi) else 'N/A',
            }

            for idx, (key, value) in enumerate(stats.items()):
                ws[f'A{stats_row + idx + 1}'] = key
                ws[f'B{stats_row + idx + 1}'] = value
                ws[f'A{stats_row + idx + 1}'].font = Font(name=self.base_font_name, bold=True)
            stats_last_row = stats_row + len(stats)
            self._add_thin_border_box(ws, stats_row, stats_last_row, 1, 2)

        if include_charts and len(df_out) > 1 and 'Close' in col_index:
            try:
                chart = LineChart()
                chart.title = f"{ticker} Price with Moving Averages"
                chart.style = 13
                chart.y_axis.title = 'Price ($)'
                chart.x_axis.title = 'Date'
                chart.height = 10
                chart.width = 22

                close_col = col_index['Close']
                max_chart_row = min(last_data_row, 503)
                dates = Reference(ws, min_col=1, min_row=4, max_row=max_chart_row)

                data = Reference(ws, min_col=close_col, min_row=3, max_row=max_chart_row)
                chart.add_data(data, titles_from_data=True)

                for ma_col in ['SMA 20', 'SMA 50']:
                    if ma_col in col_index:
                        ma_ref = Reference(ws, min_col=col_index[ma_col], min_row=3, max_row=max_chart_row)
                        chart.add_data(ma_ref, titles_from_data=True)

                chart.set_categories(dates)
                ws.add_chart(chart, f'{get_column_letter(n_cols + 2)}5')

                if 'RSI (14)' in col_index:
                    rsi_chart = LineChart()
                    rsi_chart.title = f"{ticker} RSI (14)"
                    rsi_chart.style = 13
                    rsi_chart.y_axis.title = 'RSI'
                    rsi_chart.y_axis.scaling.min = 0
                    rsi_chart.y_axis.scaling.max = 100
                    rsi_chart.height = 7
                    rsi_chart.width = 22
                    rsi_ref = Reference(ws, min_col=col_index['RSI (14)'], min_row=3, max_row=max_chart_row)
                    rsi_chart.add_data(rsi_ref, titles_from_data=True)
                    rsi_chart.set_categories(dates)
                    ws.add_chart(rsi_chart, f'{get_column_letter(n_cols + 2)}26')
            except Exception:
                pass

        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols)

    def _create_analysis_sheet(self, writer, sheet_name, df):
        """Create formatted analysis sheet (Metric/Value style tables)"""
        clean_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=clean_name, startrow=2)

        ws = writer.sheets[clean_name]
        self._set_tab_color(ws, 'analysis')
        ws.sheet_view.showGridLines = False
        self._register_sheet(clean_name, sheet_name.replace('_', ' ').title(), 'analysis',
                              "Analytics and calculated metrics")

        n_cols = len(df.columns) + 1  # +1 for index column

        ws['A1'] = sheet_name.replace('_', ' ').title()
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['accent'], end_color=self.colors['accent'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols - 1, 1))}1')
        self._add_nav_link(ws, n_cols)

        for idx, col in enumerate(['Index'] + list(df.columns)):
            cell = ws.cell(row=3, column=idx + 1)
            cell.value = col if idx > 0 else ''
        self._style_header_row(ws, 3, n_cols)

        self._autosize_columns(ws, n_cols, start_row=1, end_row=ws.max_row, min_width=10, max_width=25)
        self._apply_zebra(ws, 4, ws.max_row, 1, n_cols)
        self._add_thin_border_box(ws, 3, ws.max_row, 1, n_cols)
        self._center_all_cells(ws)
        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols)

    def _create_comparison_sheet(self, writer, ticker_data, metadata, include_charts=True):
        """Create comparison sheet for multiple tickers, plus a 'Growth of $100' rebased
        performance chart so tickers with very different price levels can be compared visually."""
        comparison_data = []

        for ticker, df in ticker_data.items():
            if len(df) > 1 and 'Close' in df.columns:
                returns = df['Close'].pct_change()

                stats = {
                    'Ticker': ticker,
                    'Start Date': df.index[0].strftime('%Y-%m-%d'),
                    'End Date': df.index[-1].strftime('%Y-%m-%d'),
                    'Days': len(df),
                    'Start Price': df['Close'].iloc[0],
                    'End Price': df['Close'].iloc[-1],
                    'Total Return (%)': ((df['Close'].iloc[-1] / df['Close'].iloc[0] - 1) * 100),
                    'Annualized Return (%)': (((df['Close'].iloc[-1] / df['Close'].iloc[0]) ** (252 / len(df))) - 1) * 100,
                    'Volatility (%)': returns.std() * np.sqrt(252) * 100,
                    'Max Drawdown (%)': self._calculate_max_drawdown(df['Close']),
                    'Sharpe Ratio': (returns.mean() / returns.std()) * np.sqrt(252) if returns.std() > 0 else 0,
                    'Best Day (%)': returns.max() * 100,
                    'Worst Day (%)': returns.min() * 100,
                    'Positive Days (%)': (returns > 0).sum() / len(returns.dropna()) * 100,
                    'Avg Volume': int(df['Volume'].mean()) if 'Volume' in df.columns else 0
                }
                comparison_data.append(stats)

        if not comparison_data:
            return

        df_comp = pd.DataFrame(comparison_data)
        df_comp.to_excel(writer, sheet_name='Comparison', index=False, startrow=2)

        ws = writer.sheets['Comparison']
        self._set_tab_color(ws, 'comparison')
        ws.sheet_view.showGridLines = False
        self._register_sheet('Comparison', 'Multi-Ticker Comparison', 'comparison',
                              "Side-by-side risk/return stats plus a rebased 'Growth of $100' chart")

        n_cols = len(df_comp.columns)
        ws['A1'] = 'Multi-Ticker Comparison Analysis'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols - 2, 1))}1')
        self._add_nav_link(ws, n_cols)

        for idx, col in enumerate(df_comp.columns):
            ws.cell(row=3, column=idx + 1, value=col)
        self._style_header_row(ws, 3, n_cols)

        first_row, last_row = 4, len(df_comp) + 3
        col_index = {name: i + 1 for i, name in enumerate(df_comp.columns)}

        for money_col in ['Start Price', 'End Price']:
            self._format_numeric_column(ws, col_index[money_col], first_row, last_row, '$#,##0.00')
        for pct_col in ['Total Return (%)', 'Annualized Return (%)', 'Volatility (%)',
                        'Max Drawdown (%)', 'Best Day (%)', 'Worst Day (%)', 'Positive Days (%)']:
            c = col_index[pct_col]
            for r in range(first_row, last_row + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100
                    cell.number_format = '+0.00%;-0.00%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        self._format_numeric_column(ws, col_index['Sharpe Ratio'], first_row, last_row, '0.00')
        self._format_numeric_column(ws, col_index['Avg Volume'], first_row, last_row, '#,##0')

        self._apply_zebra(ws, first_row, last_row, 1, n_cols)
        self._autosize_columns(ws, n_cols, start_row=3, end_row=last_row, min_width=10, max_width=20)
        self._add_thin_border_box(ws, 3, last_row, 1, n_cols)

        # Color scale + icon set on the metrics that matter most for at-a-glance ranking
        ret_letter = get_column_letter(col_index['Total Return (%)'])
        sharpe_letter = get_column_letter(col_index['Sharpe Ratio'])
        self._add_color_scale(ws, f'{ret_letter}{first_row}:{ret_letter}{last_row}')
        self._add_icon_set(ws, f'{ret_letter}{first_row}:{ret_letter}{last_row}')
        self._add_color_scale(ws, f'{sharpe_letter}{first_row}:{sharpe_letter}{last_row}')

        table_ref = f"A3:{get_column_letter(n_cols)}{last_row}"
        self._add_excel_table(ws, table_ref, "Comparison")

        # ---- Growth of $100 rebased chart ----
        chart_data_start_row = last_row + 4
        try:
            rebased = pd.DataFrame({t: (df['Close'] / df['Close'].iloc[0]) * 100
                                     for t, df in ticker_data.items() if 'Close' in df.columns and len(df) > 1})
            rebased = rebased.dropna(how='all')
            if not rebased.empty:
                ws[f'A{chart_data_start_row - 1}'] = 'Growth of $100 (rebased, for charting below)'
                ws[f'A{chart_data_start_row - 1}'].font = Font(name=self.base_font_name, bold=True, size=11)

                rebased_reset = rebased.reset_index()
                rebased_reset.columns = ['Date'] + list(rebased.columns)
                for j, col in enumerate(rebased_reset.columns):
                    ws.cell(row=chart_data_start_row, column=j + 1, value=col)
                self._style_header_row(ws, chart_data_start_row, len(rebased_reset.columns))
                for i, row_vals in enumerate(rebased_reset.itertuples(index=False), 1):
                    r = chart_data_start_row + i
                    ws.cell(row=r, column=1, value=row_vals[0].strftime('%Y-%m-%d') if hasattr(row_vals[0], 'strftime') else str(row_vals[0]))
                    for j, val in enumerate(row_vals[1:], 2):
                        cell = ws.cell(row=r, column=j, value=round(float(val), 2) if pd.notna(val) else None)
                        cell.number_format = '#,##0.00'
                data_last_row = chart_data_start_row + len(rebased_reset)

                if include_charts:
                    growth_chart = LineChart()
                    growth_chart.title = "Growth of $100 — Rebased Performance"
                    growth_chart.style = 12
                    growth_chart.y_axis.title = 'Value of $100 invested'
                    growth_chart.x_axis.title = 'Date'
                    growth_chart.height = 11
                    growth_chart.width = 24
                    cats = Reference(ws, min_col=1, min_row=chart_data_start_row + 1, max_row=data_last_row)
                    for j in range(2, len(rebased_reset.columns) + 1):
                        data_ref = Reference(ws, min_col=j, min_row=chart_data_start_row, max_row=data_last_row)
                        growth_chart.add_data(data_ref, titles_from_data=True)
                    growth_chart.set_categories(cats)
                    ws.add_chart(growth_chart, f'{get_column_letter(n_cols + 2)}5')
        except Exception:
            pass

        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols)

    def _create_correlation_sheet(self, writer, ticker_data):
        """Correlation matrix of daily returns across tickers, with a color-scale heatmap —
        useful for spotting diversification (or lack of it) at a glance."""
        closes = pd.DataFrame({t: df['Close'] for t, df in ticker_data.items() if 'Close' in df.columns})
        if closes.shape[1] < 2:
            return
        corr = closes.pct_change().corr().round(3)

        sheet_name = 'Correlation'
        corr.to_excel(writer, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]
        self._set_tab_color(ws, 'correlation')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, 'Return Correlation Matrix', 'correlation',
                              "Pairwise correlation of daily returns across all tickers")

        n_cols = len(corr.columns) + 1
        ws['A1'] = 'Daily Return Correlation Matrix'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['secondary'], end_color=self.colors['secondary'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols - 1, 1))}1')
        self._add_nav_link(ws, n_cols)

        self._style_header_row(ws, 3, n_cols)
        first_row, last_row = 4, 3 + len(corr)
        for c in range(2, n_cols + 1):
            self._format_numeric_column(ws, c, first_row, last_row, '0.00')
            col_letter = get_column_letter(c)
            self._add_color_scale(ws, f'{col_letter}{first_row}:{col_letter}{last_row}',
                                   colors=('F8696B', 'FFFFFF', '63BE7B'))
        self._autosize_columns(ws, n_cols, start_row=3, end_row=last_row, min_width=10, max_width=14)
        self._add_thin_border_box(ws, 3, last_row, 1, n_cols)
        self._center_all_cells(ws)

        note_row = last_row + 2
        ws[f'A{note_row}'] = ("Values near +1 move together (little diversification benefit); "
                               "values near 0 or negative move independently or inversely.")
        ws[f'A{note_row}'].font = Font(name=self.base_font_name, size=9, italic=True, color='808080')
        ws.merge_cells(f'A{note_row}:{get_column_letter(n_cols)}{note_row}')

        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols)

    def _create_metrics_correlation_sheet(self, writer, ticker, metrics_df):
        """
        Sheet with earnings as the first row and fundamental metrics (P/E, EV/EBITDA,
        ROE, FCF, etc.) as subsequent rows, one column per fiscal quarter. Adds a
        scatter chart + linear trendline (R^2, equation) per metric vs earnings,
        a ranking table by R^2, and (if available) current analyst estimates.
        """
        sheet_name = f"{ticker}_Earnings_Metrics"[:31]
        metrics_df.to_excel(writer, sheet_name=sheet_name, startrow=2)

        ws = writer.sheets[sheet_name]
        self._set_tab_color(ws, 'metrics')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, f"{ticker} — Earnings vs Metrics", 'metrics',
                              "Quarterly earnings vs fundamentals, R² ranking, analyst estimates")

        n_rows = len(metrics_df)
        n_cols = len(metrics_df.columns)

        ws['A1'] = f'{ticker} — Earnings vs Fundamental Metrics'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['accent'], end_color=self.colors['accent'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols, 1))}1')
        self._add_nav_link(ws, n_cols + 1)

        self._style_header_row(ws, 3, n_cols + 1)

        earnings_row_num = 4
        for c in range(1, n_cols + 2):
            ws.cell(row=earnings_row_num, column=c).font = Font(name=self.base_font_name, bold=True)
            ws.cell(row=earnings_row_num, column=c).fill = PatternFill(
                start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid')

        # Format numeric quarterly cells to 2 decimal places for readability
        for r in range(earnings_row_num, earnings_row_num + n_rows):
            for c in range(2, n_cols + 2):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')

        self._apply_zebra(ws, earnings_row_num + 1, earnings_row_num + n_rows - 1, 1, n_cols + 1)
        self._add_thin_border_box(ws, 3, earnings_row_num + n_rows - 1, 1, n_cols + 1)

        ws.column_dimensions['A'].width = 30
        for c in range(2, n_cols + 2):
            ws.column_dimensions[get_column_letter(c)].width = 14

        # ---- R^2 of each metric vs earnings ----
        earnings_values = pd.to_numeric(metrics_df.iloc[0], errors='coerce').values.astype(float)
        r_squared = {}
        for i in range(1, n_rows):
            metric_name = metrics_df.index[i]
            metric_values = pd.to_numeric(metrics_df.iloc[i], errors='coerce').values.astype(float)
            mask = ~np.isnan(earnings_values) & ~np.isnan(metric_values)
            if mask.sum() >= 3 and np.std(metric_values[mask]) > 0 and np.std(earnings_values[mask]) > 0:
                r = np.corrcoef(earnings_values[mask], metric_values[mask])[0, 1]
                r_squared[metric_name] = r ** 2
            else:
                r_squared[metric_name] = np.nan

        summary_start_row = n_rows + 6
        ws[f'A{summary_start_row}'] = 'Correlation Strength with Earnings (R²)'
        ws[f'A{summary_start_row}'].font = Font(name=self.base_font_name, bold=True, size=12)

        ws[f'A{summary_start_row + 1}'] = 'Metric'
        ws[f'B{summary_start_row + 1}'] = 'R²'
        self._style_header_row(ws, summary_start_row + 1, 2)

        ranked = sorted(r_squared.items(), key=lambda x: (x[1] if pd.notna(x[1]) else -1), reverse=True)
        r2_first_row = summary_start_row + 2
        for i, (metric, r2) in enumerate(ranked):
            row = r2_first_row + i
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = round(float(r2), 4) if pd.notna(r2) else None
            if pd.notna(r2):
                ws[f'B{row}'].number_format = '0.0000'
        r2_last_row = r2_first_row + len(ranked) - 1
        if len(ranked) > 0:
            self._add_color_scale(ws, f'B{r2_first_row}:B{r2_last_row}')
            self._apply_zebra(ws, r2_first_row, r2_last_row, 1, 2)
            self._add_thin_border_box(ws, summary_start_row + 1, r2_last_row, 1, 2)
            self._add_excel_table(ws, f'A{summary_start_row + 1}:B{r2_last_row}', f"{ticker}_R2")

        # ---- Analyst estimates / current snapshot ----
        snapshot = metrics_df.attrs.get('snapshot', {})
        if snapshot:
            snap_row = r2_last_row + 3 if len(ranked) > 0 else summary_start_row + 3
            ws[f'A{snap_row}'] = 'Current Snapshot & Analyst Estimates'
            ws[f'A{snap_row}'].font = Font(name=self.base_font_name, bold=True, size=12)
            ws[f'A{snap_row + 1}'] = 'Metric'
            ws[f'B{snap_row + 1}'] = 'Value'
            self._style_header_row(ws, snap_row + 1, 2)
            for i, (key, val) in enumerate(snapshot.items()):
                row = snap_row + 2 + i
                ws[f'A{row}'] = key
                ws[f'B{row}'] = val if val is not None else 'N/A'
            self._apply_zebra(ws, snap_row + 2, snap_row + 1 + len(snapshot), 1, 2)
            self._add_thin_border_box(ws, snap_row + 1, snap_row + 1 + len(snapshot), 1, 2)

        # ---- Scatter chart + linear trendline per metric ----
        chart_col = get_column_letter(n_cols + 3)
        chart_row_cursor = 4

        for i in range(1, n_rows):
            metric_name = metrics_df.index[i]
            data_row = i + 4

            try:
                chart = ScatterChart()
                chart.title = f"{metric_name} vs Earnings"
                chart.style = 13
                chart.x_axis.title = 'Earnings (Net Income)'
                chart.y_axis.title = str(metric_name)
                chart.height = 8
                chart.width = 15

                xvalues = Reference(ws, min_col=2, max_col=n_cols + 1,
                                     min_row=earnings_row_num, max_row=earnings_row_num)
                yvalues = Reference(ws, min_col=2, max_col=n_cols + 1,
                                     min_row=data_row, max_row=data_row)

                series = Series(yvalues, xvalues, title=str(metric_name))
                series.marker.symbol = 'circle'
                series.graphicalProperties.line.noFill = True
                series.trendline = Trendline(trendlineType='linear', dispRSqr=True, dispEq=True)

                chart.series.append(series)
                ws.add_chart(chart, f'{chart_col}{chart_row_cursor}')
                chart_row_cursor += 17
            except Exception:
                pass

        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols + 1)

    def _create_dividends_sheet(self, writer, ticker, divs, splits):
        """Dividend payment history and stock-split history for a ticker."""
        sheet_name = f"{ticker}_Dividends"[:31]

        div_df = pd.DataFrame({
            'Date': divs.index.strftime('%Y-%m-%d') if divs is not None and not divs.empty else [],
            'Dividend': divs.values if divs is not None and not divs.empty else []
        })
        split_df = pd.DataFrame({
            'Date': splits.index.strftime('%Y-%m-%d') if splits is not None and not splits.empty else [],
            'Split Ratio': splits.values if splits is not None and not splits.empty else []
        })

        div_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ws = writer.sheets[sheet_name]
        self._set_tab_color(ws, 'dividends')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, f"{ticker} — Dividends & Splits", 'dividends',
                              "Full dividend payment and stock-split history")

        ws['A1'] = f'{ticker} — Dividends & Stock Splits'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['primary'], end_color=self.colors['primary'], fill_type='solid')
        ws.merge_cells('A1:B1')
        self._add_nav_link(ws, 4)

        self._style_header_row(ws, 3, 2)

        n_divs = len(div_df)
        if n_divs > 0:
            first_row, last_row = 4, 3 + n_divs
            self._format_numeric_column(ws, 2, first_row, last_row, '$#,##0.0000')
            self._apply_zebra(ws, first_row, last_row, 1, 2)
            self._add_thin_border_box(ws, 3, last_row, 1, 2)
            self._add_excel_table(ws, f'A3:B{last_row}', f"{ticker}_Dividends")

            total_paid = float(div_df['Dividend'].sum())
            annualized_recent = float(div_df['Dividend'].tail(4).sum()) if n_divs >= 1 else 0.0
            summary_row = last_row + 2
            ws[f'A{summary_row}'] = 'Total Dividends Paid (period)'
            ws[f'B{summary_row}'] = round(total_paid, 4)
            ws[f'A{summary_row + 1}'] = 'Approx. Trailing 4-Payment Total'
            ws[f'B{summary_row + 1}'] = round(annualized_recent, 4)
            ws[f'A{summary_row}'].font = Font(name=self.base_font_name, bold=True)
            ws[f'A{summary_row + 1}'].font = Font(name=self.base_font_name, bold=True)
        else:
            ws['A4'] = 'No dividend history found for this ticker.'

        # Splits placed to the right, out of the way of the dividend table
        split_start_col = 4  # column D
        ws.cell(row=3, column=split_start_col, value='Date')
        ws.cell(row=3, column=split_start_col + 1, value='Split Ratio')
        self._style_header_row(ws, 3, 2, start_col=split_start_col)
        if len(split_df) > 0:
            for i, row in split_df.iterrows():
                r = 4 + i
                ws.cell(row=r, column=split_start_col, value=row['Date'])
                ws.cell(row=r, column=split_start_col + 1, value=float(row['Split Ratio']))
            self._apply_zebra(ws, 4, 3 + len(split_df), split_start_col, split_start_col + 1)
            self._add_thin_border_box(ws, 3, 3 + len(split_df), split_start_col, split_start_col + 1)
        else:
            ws.cell(row=4, column=split_start_col, value='No stock splits in this period.')

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.freeze_panes = 'A4'
        self._apply_print_setup(ws, 5, orientation='portrait')

    def _create_benchmark_sheet(self, writer, ticker, bdata, include_charts=True):
        """
        Relative performance vs a benchmark (e.g. SPY): cumulative relative return,
        rolling beta, rolling correlation, plus a summary stats block
        (beta, alpha, correlation, tracking error, information ratio).
        """
        benchmark_ticker = bdata['benchmark_ticker']
        rel_df = bdata['rel_df']
        stats = bdata['stats']

        sheet_name = f"{ticker}_vs_{benchmark_ticker}"[:31]
        rel_df.to_excel(writer, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]
        self._set_tab_color(ws, 'benchmark')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, f"{ticker} vs {benchmark_ticker}", 'benchmark',
                              "Relative performance, rolling beta/correlation, alpha & tracking error")

        n_cols = len(rel_df.columns) + 1
        ws['A1'] = f'{ticker} vs {benchmark_ticker} — Relative Performance'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['danger'], end_color=self.colors['danger'], fill_type='solid')
        ws.merge_cells(f'A1:{get_column_letter(max(n_cols - 1, 1))}1')
        self._add_nav_link(ws, n_cols)

        for idx, col in enumerate(['Date'] + list(rel_df.columns)):
            ws.cell(row=3, column=idx + 1, value=col)
        self._style_header_row(ws, 3, n_cols)

        first_row, last_row = 4, len(rel_df) + 3
        ws.column_dimensions['A'].width = 14
        for r in range(first_row, last_row + 1):
            ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'

        col_index = {name: i + 2 for i, name in enumerate(rel_df.columns)}
        for pct_col in [c for c in rel_df.columns if 'Return' in c or 'Cumulative' in c]:
            self._format_numeric_column(ws, col_index[pct_col], first_row, last_row, '+0.00%;-0.00%')
        for other_col in [c for c in rel_df.columns if c not in col_index or ('Beta' in c or 'Corr' in c)]:
            if other_col in col_index:
                self._format_numeric_column(ws, col_index[other_col], first_row, last_row, '0.00')

        self._autosize_columns(ws, n_cols - 1, start_row=3, end_row=last_row, start_col=2, min_width=12, max_width=20)
        self._apply_zebra(ws, first_row, last_row, 1, n_cols)
        self._add_thin_border_box(ws, 3, last_row, 1, n_cols)

        # Summary stats block
        summary_row = last_row + 3
        ws[f'A{summary_row}'] = f'Summary vs {benchmark_ticker}'
        ws[f'A{summary_row}'].font = Font(name=self.base_font_name, bold=True, size=12)
        labels = ['Beta', 'Alpha (Annualized %)', 'Correlation', 'Tracking Error (Annualized %)', 'Information Ratio']
        for i, label in enumerate(labels):
            row = summary_row + 1 + i
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name=self.base_font_name, bold=True)
            val = stats.get(label)
            ws[f'B{row}'] = round(float(val), 4) if val is not None and pd.notna(val) else 'N/A'
        self._add_thin_border_box(ws, summary_row, summary_row + len(labels), 1, 2)

        if include_charts and 'Cumulative Relative Return' in rel_df.columns:
            try:
                chart = LineChart()
                chart.title = f"{ticker} vs {benchmark_ticker} — Cumulative Relative Return"
                chart.style = 13
                chart.y_axis.title = 'Cumulative Relative Return'
                chart.x_axis.title = 'Date'
                chart.height = 10
                chart.width = 20

                rel_col = col_index['Cumulative Relative Return']
                data = Reference(ws, min_col=rel_col, min_row=3, max_row=last_row)
                dates = Reference(ws, min_col=1, min_row=4, max_row=last_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(dates)
                ws.add_chart(chart, f'{get_column_letter(n_cols + 2)}5')
            except Exception:
                pass

        ws.freeze_panes = 'B4'
        self._apply_print_setup(ws, n_cols)

    def _create_seasonality_sheet(self, writer, ticker, sdata, include_charts=True):
        """Average return by calendar month and by day-of-week, with bar charts."""
        monthly = sdata.get('monthly')
        weekday = sdata.get('weekday')

        sheet_name = f"{ticker}_Seasonality"[:31]
        start_row = 2
        if monthly is not None and not monthly.empty:
            monthly.to_excel(writer, sheet_name=sheet_name, startrow=start_row)
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            # Nothing to write (no monthly data) — still create a minimal sheet
            pd.DataFrame({'Note': ['Not enough history for seasonality analysis.']}).to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=start_row)
            ws = writer.sheets[sheet_name]

        self._set_tab_color(ws, 'seasonality')
        ws.sheet_view.showGridLines = False
        self._register_sheet(sheet_name, f"{ticker} — Seasonality", 'seasonality',
                              "Average return by calendar month and day of week")

        ws['A1'] = f'{ticker} — Seasonality Analysis'
        ws['A1'].font = Font(name=self.base_font_name, size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['secondary'], end_color=self.colors['secondary'], fill_type='solid')
        ws.merge_cells('A1:C1')
        self._add_nav_link(ws, 4)

        if monthly is not None and not monthly.empty:
            n_cols_m = len(monthly.columns) + 1
            self._style_header_row(ws, start_row + 1, n_cols_m)
            m_first, m_last = start_row + 2, start_row + 1 + len(monthly)
            for c in range(2, n_cols_m + 1):
                self._format_numeric_column(ws, c, m_first, m_last, '+0.00%;-0.00%')
            self._apply_zebra(ws, m_first, m_last, 1, n_cols_m)
            self._add_thin_border_box(ws, start_row + 1, m_last, 1, n_cols_m)
            self._autosize_columns(ws, n_cols_m, start_row=start_row + 1, end_row=m_last, min_width=10, max_width=20)

            if include_charts:
                try:
                    chart = BarChart()
                    chart.title = f"{ticker} — Average Return by Month"
                    chart.y_axis.title = 'Avg Return'
                    chart.x_axis.title = 'Month'
                    chart.height = 8
                    chart.width = 16
                    data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=m_last)
                    cats = Reference(ws, min_col=1, min_row=m_first, max_row=m_last)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws.add_chart(chart, f'{get_column_letter(n_cols_m + 2)}2')
                except Exception:
                    pass

            next_block_row = m_last + 3
        else:
            next_block_row = start_row + 2

        if weekday is not None and not weekday.empty:
            ws[f'A{next_block_row}'] = 'Average Return by Day of Week'
            ws[f'A{next_block_row}'].font = Font(name=self.base_font_name, bold=True, size=12)
            wd_header_row = next_block_row + 1
            ws.cell(row=wd_header_row, column=1, value='Day')
            ws.cell(row=wd_header_row, column=2, value='Avg Return')
            self._style_header_row(ws, wd_header_row, 2)
            wd_first = wd_header_row + 1
            for i, (day, val) in enumerate(weekday.items()):
                r = wd_first + i
                ws.cell(row=r, column=1, value=str(day))
                cell = ws.cell(row=r, column=2, value=round(float(val), 6))
                cell.number_format = '+0.00%;-0.00%'
            wd_last = wd_first + len(weekday) - 1
            self._apply_zebra(ws, wd_first, wd_last, 1, 2)
            self._add_thin_border_box(ws, wd_header_row, wd_last, 1, 2)

            if include_charts:
                try:
                    chart2 = BarChart()
                    chart2.title = f"{ticker} — Average Return by Day of Week"
                    chart2.y_axis.title = 'Avg Return'
                    chart2.height = 8
                    chart2.width = 16
                    data2 = Reference(ws, min_col=2, min_row=wd_header_row, max_row=wd_last)
                    cats2 = Reference(ws, min_col=1, min_row=wd_first, max_row=wd_last)
                    chart2.add_data(data2, titles_from_data=True)
                    chart2.set_categories(cats2)
                    ws.add_chart(chart2, f'D{wd_header_row}')
                except Exception:
                    pass

        ws.freeze_panes = 'A4'
        self._apply_print_setup(ws, 4, orientation='portrait')

    def _calculate_max_drawdown(self, prices):
        """Calculate maximum drawdown"""
        cumulative = (1 + prices.pct_change()).cumprod()
        running_max = cumulative.expanding().max()
        drawdown = ((cumulative - running_max) / running_max) * 100
        return drawdown.min()


# ----------------------------------------------------------------------
# Module-level data helpers (kept free of st.* calls so they're safe to cache)
# ----------------------------------------------------------------------

def _pick_field(row: pd.Series, candidates: List[str]):
    """
    Return the first present, non-null value among several possible yfinance
    row-label spellings, since exact labels vary across versions/tickers.
    """
    for name in candidates:
        if name in row.index:
            val = row.get(name)
            if pd.notna(val):
                return val
    return np.nan


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_fundamental_metrics(ticker: str, max_quarters: int = 8) -> Optional[pd.DataFrame]:
    """
    Build a quarterly earnings-vs-metrics table for a ticker, plus a snapshot of
    current valuation and analyst-estimate fields (attached via df.attrs['snapshot']).

    Returns None if quarterly financial data isn't available (common for ETFs,
    indices, forex, futures, and some crypto pairs). Pure function (no st.* calls)
    so it is safe to cache; callers should surface a message on None themselves.
    """
    try:
        t = yf.Ticker(ticker)

        income = getattr(t, 'quarterly_income_stmt', None)
        if income is None or income.empty:
            income = t.quarterly_financials
        balance = t.quarterly_balance_sheet
        cashflow = t.quarterly_cashflow

        try:
            info = t.info or {}
        except Exception:
            info = {}

        if income is None or income.empty:
            return None

        income = income.T.sort_index()
        income = income.tail(max_quarters)
        balance = balance.T.sort_index() if balance is not None and not balance.empty else pd.DataFrame()
        cashflow = cashflow.T.sort_index() if cashflow is not None and not cashflow.empty else pd.DataFrame()

        quarters = income.index

        start = (quarters.min() - pd.Timedelta(days=10)).date()
        end = (quarters.max() + pd.Timedelta(days=10)).date()
        try:
            price_hist = t.history(start=start, end=end)
        except Exception:
            price_hist = pd.DataFrame()

        def price_near(dt):
            if price_hist.empty:
                return np.nan
            try:
                pos = price_hist.index.get_indexer([dt], method='nearest')[0]
                return float(price_hist['Close'].iloc[pos])
            except Exception:
                return np.nan

        shares_out = info.get('sharesOutstanding', np.nan)

        earnings_row, eps_row, rev_row = {}, {}, {}
        pe_row, margin_row, roe_row = {}, {}, {}
        ev_ebitda_row, fcf_row = {}, {}

        for q in quarters:
            row = income.loc[q]
            net_income = _pick_field(row, ['Net Income', 'Net Income Common Stockholders',
                                            'Net Income Continuous Operations', 'Net Income Including Noncontrolling Interests'])
            diluted_eps = _pick_field(row, ['Diluted EPS', 'Basic EPS'])
            total_rev = _pick_field(row, ['Total Revenue', 'Operating Revenue'])
            ebitda = _pick_field(row, ['EBITDA', 'Normalized EBITDA'])

            price = price_near(q)

            earnings_row[q] = net_income
            eps_row[q] = diluted_eps
            rev_row[q] = total_rev

            if pd.notna(diluted_eps) and diluted_eps != 0 and pd.notna(price):
                pe_row[q] = price / (diluted_eps * 4)
            else:
                pe_row[q] = np.nan

            if pd.notna(net_income) and pd.notna(total_rev) and total_rev not in (0, np.nan):
                margin_row[q] = (net_income / total_rev) * 100
            else:
                margin_row[q] = np.nan

            total_debt, cash = 0, 0
            if q in balance.index:
                brow = balance.loc[q]
                equity = _pick_field(brow, ['Stockholders Equity', 'Total Equity Gross Minority Interest', 'Common Stock Equity'])
                if pd.notna(equity) and equity != 0 and pd.notna(net_income):
                    roe_row[q] = (net_income / equity) * 100
                else:
                    roe_row[q] = np.nan
                total_debt = _pick_field(brow, ['Total Debt', 'Net Debt']) or 0
                cash = _pick_field(brow, ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments']) or 0
            else:
                roe_row[q] = np.nan

            if pd.notna(price) and pd.notna(shares_out) and pd.notna(ebitda) and ebitda not in (0, np.nan):
                market_cap = price * shares_out
                ev = market_cap + total_debt - cash
                ev_ebitda_row[q] = ev / (ebitda * 4)
            else:
                ev_ebitda_row[q] = np.nan

            if q in cashflow.index:
                crow = cashflow.loc[q]
                fcf = _pick_field(crow, ['Free Cash Flow'])
                if pd.isna(fcf):
                    ocf = _pick_field(crow, ['Operating Cash Flow', 'Cash Flow From Continuing Operating Activities'])
                    capex = _pick_field(crow, ['Capital Expenditure', 'Capital Expenditure Reported'])
                    if pd.notna(ocf) and pd.notna(capex):
                        fcf = ocf + capex
                fcf_row[q] = fcf
            else:
                fcf_row[q] = np.nan

        col_labels = [q.strftime('%Y-%m-%d') for q in quarters]

        data = {
            'Earnings (Net Income)': list(earnings_row.values()),
            'Diluted EPS': list(eps_row.values()),
            'Revenue': list(rev_row.values()),
            'Net Margin (%)': list(margin_row.values()),
            'P/E (qtr-annualized)': list(pe_row.values()),
            'EV/EBITDA (qtr-annualized)': list(ev_ebitda_row.values()),
            'ROE (%)': list(roe_row.values()),
            'Free Cash Flow': list(fcf_row.values()),
        }

        df = pd.DataFrame(data, index=col_labels).T

        def pct(x):
            return round(x * 100, 2) if isinstance(x, (int, float)) else np.nan

        df.attrs['snapshot'] = {
            'Trailing P/E': info.get('trailingPE'),
            'Forward P/E': info.get('forwardPE'),
            'PEG Ratio': info.get('pegRatio') or info.get('trailingPegRatio'),
            'EV/EBITDA (current)': info.get('enterpriseToEbitda'),
            'Price/Book': info.get('priceToBook'),
            'ROE (current, %)': pct(info.get('returnOnEquity')),
            'ROA (current, %)': pct(info.get('returnOnAssets')),
            'Analyst Target (Mean)': info.get('targetMeanPrice'),
            'Analyst Target (High)': info.get('targetHighPrice'),
            'Analyst Target (Low)': info.get('targetLowPrice'),
            'Analyst Recommendation': info.get('recommendationKey'),
            'Number of Analyst Opinions': info.get('numberOfAnalystOpinions'),
        }

        return df

    except Exception:
        return None


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_dividends_splits(ticker: str) -> Tuple[pd.Series, pd.Series]:
    """Fetch full dividend and stock-split history for a ticker. Pure/cacheable."""
    try:
        t = yf.Ticker(ticker)
        divs = t.dividends
        splits = t.splits
        return (divs if divs is not None else pd.Series(dtype=float),
                splits if splits is not None else pd.Series(dtype=float))
    except Exception:
        return pd.Series(dtype=float), pd.Series(dtype=float)


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_company_profile(ticker: str) -> Optional[Dict]:
    """
    Fetch a lightweight company-profile snapshot for the Summary sheet:
    sector, industry, market cap, employee count, 52-week range, beta, and
    dividend yield. Pure function (no st.* calls) so it is safe to cache.
    Returns None for instruments without a standard company profile
    (ETFs, indices, forex, futures, most crypto).
    """
    try:
        info = yf.Ticker(ticker).info or {}
        if not info or info.get('quoteType') not in (None, 'EQUITY'):
            # Still allow through if it has at least sector/industry info
            if not info.get('sector') and not info.get('longName'):
                return None
        return {
            'Company Name': info.get('longName') or info.get('shortName'),
            'Sector': info.get('sector'),
            'Industry': info.get('industry'),
            'Exchange': info.get('exchange'),
            'Currency': info.get('currency'),
            'Market Cap': info.get('marketCap'),
            'Employees': info.get('fullTimeEmployees'),
            'Website': info.get('website'),
            '52W High': info.get('fiftyTwoWeekHigh'),
            '52W Low': info.get('fiftyTwoWeekLow'),
            'Beta': info.get('beta'),
            'Dividend Yield': info.get('dividendYield'),
        }
    except Exception:
        return None


@st.cache_data(ttl=1800, show_spinner=False)
def fetch_benchmark_prices(benchmark_ticker: str, start_date, end_date, interval: str) -> Optional[pd.DataFrame]:
    """Fetch benchmark OHLCV data (e.g. SPY) for relative-performance comparison."""
    try:
        df = yf.download(benchmark_ticker, start=start_date, end=end_date,
                          interval=interval, progress=False, auto_adjust=False, actions=False)
        if df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)
        return df
    except Exception:
        return None


def calculate_benchmark_analytics(ticker: str, df: pd.DataFrame, benchmark_ticker: str,
                                   bdf: pd.DataFrame, rolling_window: int = 60) -> Optional[Dict]:
    """
    Build the relative-performance DataFrame and summary stats (beta, alpha,
    correlation, tracking error, information ratio) for ticker vs benchmark.
    """
    if 'Close' not in df.columns or 'Close' not in bdf.columns:
        return None

    aligned = pd.DataFrame({
        f'{ticker} Close': df['Close'],
        f'{benchmark_ticker} Close': bdf['Close']
    }).dropna()

    if len(aligned) < 5:
        return None

    r_t = aligned[f'{ticker} Close'].pct_change()
    r_b = aligned[f'{benchmark_ticker} Close'].pct_change()

    rel_df = pd.DataFrame(index=aligned.index)
    rel_df[f'{ticker} Return'] = r_t
    rel_df[f'{benchmark_ticker} Return'] = r_b
    rel_df['Relative Return'] = r_t - r_b
    rel_df['Cumulative Relative Return'] = (1 + rel_df['Relative Return'].fillna(0)).cumprod() - 1

    win = min(rolling_window, max(len(aligned) // 3, 5))
    rolling_cov = r_t.rolling(win).cov(r_b)
    rolling_var = r_b.rolling(win).var()
    rel_df[f'Rolling Beta ({win}d)'] = rolling_cov / rolling_var
    rel_df[f'Rolling Correlation ({win}d)'] = r_t.rolling(win).corr(r_b)

    valid = pd.DataFrame({'rt': r_t, 'rb': r_b}).dropna()
    if len(valid) >= 5 and valid['rb'].std() > 0:
        beta = valid['rt'].cov(valid['rb']) / valid['rb'].var()
        ann_t = valid['rt'].mean() * 252
        ann_b = valid['rb'].mean() * 252
        alpha = (ann_t - beta * ann_b) * 100
        corr = valid['rt'].corr(valid['rb'])
        tracking_error = (valid['rt'] - valid['rb']).std() * np.sqrt(252) * 100
        info_ratio = ((valid['rt'] - valid['rb']).mean() * 252) / ((valid['rt'] - valid['rb']).std() * np.sqrt(252)) \
            if (valid['rt'] - valid['rb']).std() > 0 else np.nan
    else:
        beta = alpha = corr = tracking_error = info_ratio = np.nan

    stats = {
        'Beta': beta,
        'Alpha (Annualized %)': alpha,
        'Correlation': corr,
        'Tracking Error (Annualized %)': tracking_error,
        'Information Ratio': info_ratio,
    }

    rel_df.index.name = 'Date'
    return {'benchmark_ticker': benchmark_ticker, 'rel_df': rel_df, 'stats': stats}


def calculate_seasonality(df: pd.DataFrame) -> Optional[Dict[str, pd.DataFrame]]:
    """Average return by calendar month (Jan-Dec) and by day of week."""
    if 'Close' not in df.columns or len(df) < 30:
        return None

    returns = df['Close'].pct_change().dropna()
    if returns.empty:
        return None

    monthly_avg = returns.groupby(returns.index.month).mean()
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_df = pd.DataFrame({
        'Avg Daily Return': [monthly_avg.get(m, np.nan) for m in range(1, 13)]
    }, index=[month_names[m - 1] for m in range(1, 13)])

    weekday_avg = returns.groupby(returns.index.dayofweek).mean()
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    weekday_series = pd.Series(
        {day_names[d]: weekday_avg.get(d, np.nan) for d in range(7) if d in weekday_avg.index}
    )

    return {'monthly': monthly_df, 'weekday': weekday_series}


def calculate_advanced_analytics(df: pd.DataFrame, ticker: str) -> Dict[str, pd.DataFrame]:
    """
    Calculate comprehensive analytics from price data.
    Returns dict of DataFrames for different analysis sheets.
    """
    analytics = {}

    if 'Close' in df.columns and len(df) > 1:
        returns = df['Close'].pct_change()

        returns_data = {
            'Metric': [
                'Total Return', 'Annualized Return', 'Annualized Volatility',
                'Sharpe Ratio (Rf=2%)', 'Sortino Ratio', 'Max Drawdown', 'Calmar Ratio',
                'Best Day', 'Worst Day', 'Best Month', 'Worst Month', 'Positive Days',
                'Average Daily Return', 'Average Positive Day', 'Average Negative Day',
                'Win Rate', 'Profit Factor', 'Current Price', 'Period High', 'Period Low',
                'Distance from High', 'Distance from Low'
            ],
            'Value': []
        }

        total_return = (df['Close'].iloc[-1] / df['Close'].iloc[0] - 1) * 100
        days = len(df)
        ann_return = ((1 + total_return / 100) ** (252 / days) - 1) * 100
        volatility = returns.std() * np.sqrt(252) * 100

        excess_returns = returns - (0.02 / 252)
        sharpe = (excess_returns.mean() / returns.std()) * np.sqrt(252) if returns.std() > 0 else 0

        downside_returns = returns[returns < 0]
        downside_std = downside_returns.std() * np.sqrt(252)
        sortino = (ann_return - 2) / downside_std if downside_std > 0 else 0

        cum_returns = (1 + returns).cumprod()
        running_max = cum_returns.expanding().max()
        drawdown = ((cum_returns - running_max) / running_max).min() * 100

        calmar = ann_return / abs(drawdown) if drawdown != 0 else 0

        best_day = returns.max() * 100
        worst_day = returns.min() * 100

        df_monthly = df.copy()
        df_monthly['Month'] = df_monthly.index.to_period('M')
        monthly_returns = df_monthly.groupby('Month')['Close'].apply(lambda x: (x.iloc[-1] / x.iloc[0] - 1) * 100)
        best_month = monthly_returns.max() if len(monthly_returns) > 0 else 0
        worst_month = monthly_returns.min() if len(monthly_returns) > 0 else 0

        positive_days = (returns > 0).sum()
        total_days = len(returns.dropna())
        win_rate = (positive_days / total_days * 100) if total_days > 0 else 0

        avg_positive = returns[returns > 0].mean() * 100 if len(returns[returns > 0]) > 0 else 0
        avg_negative = returns[returns < 0].mean() * 100 if len(returns[returns < 0]) > 0 else 0

        profit_factor = abs(returns[returns > 0].sum() / returns[returns < 0].sum()) if returns[returns < 0].sum() != 0 else 0

        current = df['Close'].iloc[-1]
        high = df['High'].max() if 'High' in df.columns else df['Close'].max()
        low = df['Low'].min() if 'Low' in df.columns else df['Close'].min()

        # VaR (95%, historical) — a common "how bad could a normal day get" risk metric
        var_95 = np.percentile(returns.dropna(), 5) * 100 if len(returns.dropna()) > 0 else 0

        returns_data['Metric'].append('Value at Risk (95%, daily)')
        returns_data['Value'] = [
            f"{total_return:.2f}%", f"{ann_return:.2f}%", f"{volatility:.2f}%",
            f"{sharpe:.2f}", f"{sortino:.2f}", f"{drawdown:.2f}%", f"{calmar:.2f}",
            f"{best_day:.2f}%", f"{worst_day:.2f}%", f"{best_month:.2f}%", f"{worst_month:.2f}%",
            f"{positive_days} ({win_rate:.1f}%)", f"{returns.mean() * 100:.4f}%",
            f"{avg_positive:.4f}%", f"{avg_negative:.4f}%", f"{win_rate:.2f}%", f"{profit_factor:.2f}",
            f"${current:.2f}", f"${high:.2f}", f"${low:.2f}",
            f"{((current - high) / high * 100):.2f}%", f"{((current - low) / low * 100):.2f}%",
            f"{var_95:.2f}%"
        ]

        analytics['Returns_Analysis'] = pd.DataFrame(returns_data)

    if len(df) > 30:
        df_monthly = df.copy()
        df_monthly['Month'] = df_monthly.index.to_period('M')

        monthly_stats = []
        for month, group in df_monthly.groupby('Month'):
            if len(group) > 0 and 'Close' in group.columns:
                month_return = (group['Close'].iloc[-1] / group['Close'].iloc[0] - 1) * 100
                monthly_stats.append({
                    'Month': str(month),
                    'Start': round(group['Close'].iloc[0], 2),
                    'End': round(group['Close'].iloc[-1], 2),
                    'Return (%)': round(month_return, 2),
                    'High': round(group['High'].max(), 2) if 'High' in group.columns else None,
                    'Low': round(group['Low'].min(), 2) if 'Low' in group.columns else None,
                    'Avg Volume': int(group['Volume'].mean()) if 'Volume' in group.columns else None,
                    'Volatility (%)': round(group['Close'].pct_change().std() * np.sqrt(21) * 100, 2)
                })

        if monthly_stats:
            analytics['Monthly_Performance'] = pd.DataFrame(monthly_stats)

    if 'Close' in df.columns and len(df) > 1:
        returns = df['Close'].pct_change()
        cum_returns = (1 + returns).cumprod()
        running_max = cum_returns.expanding().max()
        drawdown_series = ((cum_returns - running_max) / running_max) * 100

        drawdowns = []
        in_drawdown = False
        drawdown_start = None
        drawdown_peak = None

        for i, (dt, dd) in enumerate(drawdown_series.items()):
            if dd < -0.1 and not in_drawdown:
                in_drawdown = True
                drawdown_start = dt
                drawdown_peak = cum_returns.iloc[:i + 1].idxmax()
            elif dd >= -0.05 and in_drawdown:
                in_drawdown = False
                drawdown_trough = dt
                max_dd = drawdown_series[drawdown_peak:drawdown_trough].min()

                drawdowns.append({
                    'Peak Date': drawdown_peak.strftime('%Y-%m-%d'),
                    'Trough Date': drawdown_trough.strftime('%Y-%m-%d'),
                    'Max Drawdown (%)': round(max_dd, 2),
                    'Duration (Days)': (drawdown_trough - drawdown_peak).days,
                    'Peak Price': round(df.loc[drawdown_peak, 'Close'], 2),
                    'Trough Price': round(df.loc[drawdown_trough, 'Close'], 2)
                })

        if drawdowns:
            drawdowns_sorted = sorted(drawdowns, key=lambda x: x['Max Drawdown (%)'])[:5]
            analytics['Top_Drawdowns'] = pd.DataFrame(drawdowns_sorted)

    if 'Volume' in df.columns:
        volume_data = {
            'Metric': [
                'Average Daily Volume', 'Median Daily Volume', 'Max Volume Day', 'Min Volume Day',
                'Volume Std Dev', 'Recent 20-Day Avg', 'Volume vs 20-Day Avg',
                'High Volume Days (>1.5x avg)', 'Low Volume Days (<0.5x avg)'
            ],
            'Value': []
        }

        avg_vol = df['Volume'].mean()
        recent_avg = df['Volume'].tail(20).mean()

        volume_data['Value'] = [
            f"{int(avg_vol):,}", f"{int(df['Volume'].median()):,}", f"{int(df['Volume'].max()):,}",
            f"{int(df['Volume'].min()):,}", f"{int(df['Volume'].std()):,}", f"{int(recent_avg):,}",
            f"{((recent_avg / avg_vol - 1) * 100):+.1f}%",
            f"{(df['Volume'] > avg_vol * 1.5).sum()}", f"{(df['Volume'] < avg_vol * 0.5).sum()}"
        ]

        analytics['Volume_Analysis'] = pd.DataFrame(volume_data)

    if 'Close' in df.columns:
        closes = df['Close']

        sma_20 = closes.rolling(20).mean().iloc[-1] if len(closes) >= 20 else None
        sma_50 = closes.rolling(50).mean().iloc[-1] if len(closes) >= 50 else None
        sma_200 = closes.rolling(200).mean().iloc[-1] if len(closes) >= 200 else None

        current = closes.iloc[-1]

        tech_data = {
            'Indicator': [
                'Current Price', 'SMA 20', 'Distance from SMA 20', 'SMA 50', 'Distance from SMA 50',
                'SMA 200', 'Distance from SMA 200', '52-Week High', 'Distance from 52W High',
                '52-Week Low', 'Distance from 52W Low', 'ATR (14-day)',
                'Trend (SMA 20 vs 50)', 'Trend (SMA 50 vs 200)'
            ],
            'Value': []
        }

        if all(col in df.columns for col in ['High', 'Low', 'Close']) and len(df) >= 14:
            high_low = df['High'] - df['Low']
            high_close = abs(df['High'] - df['Close'].shift())
            low_close = abs(df['Low'] - df['Close'].shift())
            tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
            atr = tr.rolling(14).mean().iloc[-1]
        else:
            atr = None

        high_52w = closes.tail(252).max() if len(closes) >= 252 else closes.max()
        low_52w = closes.tail(252).min() if len(closes) >= 252 else closes.min()

        tech_data['Value'] = [
            f"${current:.2f}",
            f"${sma_20:.2f}" if sma_20 else "N/A",
            f"{((current / sma_20 - 1) * 100):+.2f}%" if sma_20 else "N/A",
            f"${sma_50:.2f}" if sma_50 else "N/A",
            f"{((current / sma_50 - 1) * 100):+.2f}%" if sma_50 else "N/A",
            f"${sma_200:.2f}" if sma_200 else "N/A",
            f"{((current / sma_200 - 1) * 100):+.2f}%" if sma_200 else "N/A",
            f"${high_52w:.2f}",
            f"{((current / high_52w - 1) * 100):+.2f}%",
            f"${low_52w:.2f}",
            f"{((current / low_52w - 1) * 100):+.2f}%",
            f"${atr:.2f}" if atr else "N/A",
            "Bullish" if (sma_20 and sma_50 and sma_20 > sma_50) else "Bearish" if (sma_20 and sma_50) else "N/A",
            "Bullish" if (sma_50 and sma_200 and sma_50 > sma_200) else "Bearish" if (sma_50 and sma_200) else "N/A"
        ]

        analytics['Technical_Levels'] = pd.DataFrame(tech_data)

    return analytics


def _download_ticker(ticker, start_date, end_date, interval):
    """Worker function for parallel downloads — no st.* calls allowed in threads."""
    try:
        df = yf.download(ticker, start=start_date, end=end_date, interval=interval,
                          progress=False, auto_adjust=False, actions=False)
        if df.empty:
            return ticker, None, "No data returned"

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)

        required_cols = ["Open", "High", "Low", "Close", "Adj Close", "Volume"]
        available_cols = [col for col in required_cols if col in df.columns]
        if not available_cols:
            return ticker, None, "Missing required columns"

        df = df[available_cols]
        df.index.name = "Date"

        numeric_cols = [col for col in df.columns if col != "Volume"]
        df[numeric_cols] = df[numeric_cols].round(4)
        if "Volume" in df.columns:
            df["Volume"] = df["Volume"].astype(int)

        return ticker, df, None
    except Exception as e:
        return ticker, None, str(e)


def render_excel_export() -> None:
    """
    Enhanced Excel Export module with multi-ticker support, advanced analytics,
    dividends/splits history, benchmark comparison, seasonality analysis, a
    hyperlinked table of contents, company profile snapshot, and moving
    averages/RSI on the price chart.
    """
    st.title("📥 Advanced Excel Export")
    st.markdown("Export historical stock data with professional formatting and comprehensive analytics")

    if not OPENPYXL_AVAILABLE:
        st.error("❌ openpyxl is required for Excel export. Install with: `pip install openpyxl`")
        st.info("You can still use basic export functionality, but advanced formatting will be limited.")

    st.divider()

    col1, col2 = st.columns([3, 2])
    with col1:
        export_mode = st.radio(
            "📊 Export Mode",
            options=["Single Ticker", "Multiple Tickers", "Portfolio Analysis"],
            horizontal=True,
            help="Single: One ticker with full analytics | Multiple: Batch export | Portfolio: Comparison analysis"
        )
    with col2:
        include_charts = st.checkbox("Include Charts", value=True, help="Add price charts to sheets (requires openpyxl)")

    st.divider()

    col1, col2 = st.columns([2, 1])
    with col1:
        if export_mode == "Single Ticker":
            ticker_input = st.text_input(
                "🎯 Ticker Symbol", value="AAPL",
                placeholder="Enter ticker (e.g., AAPL, MSFT, TSLA)",
                help="Examples: AAPL (stock), GC=F (gold), BTC-USD (crypto), EURUSD=X (forex), THYAO.IS (international)"
            ).strip().upper()
            tickers = [ticker_input] if ticker_input else []
        else:
            ticker_input = st.text_area(
                "📝 Ticker Symbols (one per line or comma-separated)",
                value="AAPL\nMSFT\nGOOGL\nAMZN\nTSLA", height=120,
                help="Enter multiple tickers separated by newlines or commas"
            )
            raw_tickers = ticker_input.replace(',', '\n').replace(';', '\n').split('\n')
            tickers = [t.strip().upper() for t in raw_tickers if t.strip()]

            if len(tickers) > 50:
                st.warning(f"⚠️ Limiting to first 50 tickers (you entered {len(tickers)})")
                tickers = tickers[:50]

            if tickers:
                st.success(f"✅ {len(tickers)} ticker(s) ready for export")

    with col2:
        frequency = st.selectbox("📈 Data Frequency", options=["Daily", "Weekly", "Monthly"], index=0,
                                  help="Daily = business days only | Weekly/Monthly = aggregated periods")
        include_analytics = st.checkbox("📊 Include Analytics", value=True,
                                         help="Add returns analysis, monthly stats, technical levels, and more")
        advanced_analytics = st.checkbox("🔬 Advanced Analytics", value=False,
                                          help="Include drawdown analysis, volume patterns, and technical indicators (slower)")
        include_earnings_metrics = st.checkbox(
            "💰 Earnings vs Metrics Correlation", value=True,
            help="Add a tab per ticker with earnings + fundamentals and analyst estimates, plus scatter/trendline charts. "
                 "Not available for ETFs, indices, forex, or crypto."
        )
        include_dividends = st.checkbox(
            "💵 Dividends & Stock Splits", value=True,
            help="Add a tab per ticker with full dividend payment and stock-split history."
        )
        include_seasonality = st.checkbox(
            "📆 Seasonality Analysis", value=True,
            help="Add a tab per ticker showing average return by calendar month and day of week."
        )
        include_profile = st.checkbox(
            "🏢 Company Profile Snapshot", value=True,
            help="Add sector, industry, market cap, employees, 52-week range, and beta to the Summary sheet."
        )

    st.divider()
    col1, col2 = st.columns([1, 2])
    with col1:
        compare_benchmark = st.checkbox("📈 Compare vs Benchmark", value=True,
                                         help="Adds a relative-performance tab (beta, alpha, rolling correlation) per ticker.")
    with col2:
        benchmark_ticker = st.text_input("Benchmark Ticker", value="SPY", disabled=not compare_benchmark,
                                          help="Common choices: SPY (S&P 500), QQQ (Nasdaq 100), DIA (Dow), IWM (Russell 2000)").strip().upper()

    st.subheader("📅 Date Range")
    col1, col2, col3 = st.columns(3)
    with col1:
        preset = st.selectbox("Quick Select",
                               ["Custom", "1 Month", "3 Months", "6 Months", "1 Year", "2 Years", "5 Years", "10 Years", "Max"])

    if preset != "Custom":
        end_date = date.today()
        preset_days = {
            "1 Month": 30, "3 Months": 90, "6 Months": 180, "1 Year": 365,
            "2 Years": 730, "5 Years": 1825, "10 Years": 3650
        }
        if preset in preset_days:
            start_date = end_date - timedelta(days=preset_days[preset])
        else:  # Max
            start_date = date(2000, 1, 1)

        with col2:
            start_date = st.date_input("Start Date", value=start_date, max_value=date.today())
        with col3:
            end_date = st.date_input("End Date", value=end_date, min_value=start_date, max_value=date.today())
    else:
        with col2:
            start_date = st.date_input("Start Date", value=date(2020, 1, 1), max_value=date.today())
        with col3:
            end_date = st.date_input("End Date", value=date.today(), min_value=start_date, max_value=date.today())

    if not tickers:
        st.info("👆 Please enter at least one ticker symbol above to continue.")
        return

    if start_date >= end_date:
        st.error("❌ Start date must be before end date.")
        return

    interval_map = {"Daily": "1d", "Weekly": "1wk", "Monthly": "1mo"}
    selected_interval = interval_map[frequency]

    days_diff = (end_date - start_date).days
    if frequency == "Daily":
        expected_points = days_diff * 0.7
    elif frequency == "Weekly":
        expected_points = days_diff / 7
    else:
        expected_points = days_diff / 30

    st.info(f"📊 Period: {days_diff} days | Expected data points: ~{int(expected_points)} per ticker")

    st.divider()

    export_button = st.button("🚀 Fetch Data & Generate Excel Report", type="primary", use_container_width=True)

    if export_button:
        progress_container = st.container()
        with progress_container:
            st.markdown("### 🔄 Processing Your Request")
            progress_bar = st.progress(0, text="Downloading price data...")
            status_text = st.empty()

            ticker_data = {}
            analysis_data = {}
            metrics_data = {}
            dividends_data = {}
            benchmark_data = {}
            seasonality_data = {}
            profile_data = {}
            failed_tickers = []
            failed_metrics_tickers = []
            total_records = 0

            # ── Parallel price downloads ──
            completed = 0
            with ThreadPoolExecutor(max_workers=min(8, max(len(tickers), 1))) as executor:
                futures = {
                    executor.submit(_download_ticker, ticker, start_date, end_date, selected_interval): ticker
                    for ticker in tickers
                }
                for future in as_completed(futures):
                    ticker, df, error = future.result()
                    completed += 1
                    progress_bar.progress(completed / len(tickers), text=f"Downloaded {ticker} ({completed}/{len(tickers)})")
                    if df is None:
                        failed_tickers.append((ticker, error))
                    else:
                        ticker_data[ticker] = df
                        total_records += len(df)

            # ── Benchmark data (fetched once, reused for every ticker) ──
            benchmark_df = None
            if compare_benchmark and benchmark_ticker and ticker_data:
                status_text.info(f"📈 Fetching benchmark {benchmark_ticker}...")
                benchmark_df = fetch_benchmark_prices(benchmark_ticker, start_date, end_date, selected_interval)
                if benchmark_df is None:
                    st.warning(f"⚠️ Could not fetch benchmark data for {benchmark_ticker} — skipping relative-performance tabs.")

            # ── Per-ticker analytics / metrics / dividends / benchmark / seasonality / profile ──
            for idx, ticker in enumerate(ticker_data.keys()):
                df = ticker_data[ticker]
                status_text.info(f"📊 Analyzing {ticker}... ({idx + 1}/{len(ticker_data)})")

                if include_analytics and (export_mode == "Single Ticker" or advanced_analytics):
                    ticker_analytics = calculate_advanced_analytics(df, ticker)
                    for key, value in ticker_analytics.items():
                        analysis_data[f"{ticker}_{key}"] = value

                if include_earnings_metrics:
                    mdf = fetch_fundamental_metrics(ticker)
                    if mdf is not None and not mdf.empty:
                        metrics_data[ticker] = mdf
                    else:
                        failed_metrics_tickers.append(ticker)

                if include_dividends:
                    divs, splits = fetch_dividends_splits(ticker)
                    dividends_data[ticker] = (divs, splits)

                if include_profile:
                    prof = fetch_company_profile(ticker)
                    if prof:
                        profile_data[ticker] = prof

                if compare_benchmark and benchmark_df is not None:
                    bres = calculate_benchmark_analytics(ticker, df, benchmark_ticker, benchmark_df)
                    if bres is not None:
                        benchmark_data[ticker] = bres

                if include_seasonality:
                    sres = calculate_seasonality(df)
                    if sres is not None:
                        seasonality_data[ticker] = sres

            progress_bar.progress(1.0, text="✅ Data fetch complete!")
            status_text.empty()

        st.divider()
        st.markdown("### 📈 Fetch Results")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("✅ Successful", len(ticker_data))
        with col2:
            st.metric("❌ Failed", len(failed_tickers), delta_color="inverse")
        with col3:
            st.metric("📊 Total Records", f"{total_records:,}")
        with col4:
            avg_records = total_records // max(len(ticker_data), 1)
            st.metric("📉 Avg per Ticker", f"{avg_records:,}")

        if not ticker_data:
            st.error("❌ No data could be fetched for any ticker. Please check your ticker symbols and try again.")
            if failed_tickers:
                st.error("**Failed tickers:**")
                for ticker, reason in failed_tickers:
                    st.write(f"• {ticker}: {reason}")
            return

        if failed_tickers:
            with st.expander(f"⚠️ {len(failed_tickers)} Failed Ticker(s) - Click to expand"):
                for ticker, reason in failed_tickers:
                    st.write(f"• **{ticker}**: {reason}")

        if include_earnings_metrics:
            st.divider()
            if metrics_data:
                st.success(f"💰 Earnings/Metrics tab: built for {len(metrics_data)} of {len(ticker_data)} ticker(s) "
                           f"→ look for sheets named **TICKER_Earnings_Metrics** in the downloaded file.")
            else:
                st.error("💰 Earnings/Metrics tab: not created for any ticker "
                         "(normal for ETFs, indices, forex, futures, and most crypto).")
            if failed_metrics_tickers:
                with st.expander(f"⚠️ {len(failed_metrics_tickers)} Ticker(s) without fundamentals data"):
                    for t in failed_metrics_tickers:
                        st.write(f"• {t}")

        if compare_benchmark:
            st.divider()
            if benchmark_data:
                st.success(f"📈 Benchmark comparison ({benchmark_ticker}): built for {len(benchmark_data)} of "
                           f"{len(ticker_data)} ticker(s) → sheets named **TICKER_vs_{benchmark_ticker}**.")
            elif benchmark_df is not None:
                st.warning("📈 Benchmark data was fetched but no ticker had enough overlapping history to compare.")

        if include_profile and profile_data:
            st.divider()
            st.success(f"🏢 Company profile snapshot added for {len(profile_data)} of {len(ticker_data)} ticker(s) "
                       f"→ see the **Summary** sheet.")

        if export_mode == "Portfolio Analysis" and len(ticker_data) > 1:
            st.divider()
            st.markdown("### 📊 Portfolio Overview")
            with st.spinner("Calculating portfolio metrics..."):
                combined_prices = pd.DataFrame({ticker: df['Close'] for ticker, df in ticker_data.items()})
                returns = combined_prices.pct_change()

                corr_matrix = returns.corr()
                analysis_data['Correlation_Matrix'] = corr_matrix.round(3)

                portfolio_stats = []
                for ticker in ticker_data.keys():
                    ticker_returns = returns[ticker].dropna()
                    stats = {
                        'Ticker': ticker,
                        'Total Return (%)': ((combined_prices[ticker].iloc[-1] / combined_prices[ticker].iloc[0] - 1) * 100),
                        'Ann. Return (%)': ((1 + ticker_returns.mean()) ** 252 - 1) * 100,
                        'Volatility (%)': ticker_returns.std() * np.sqrt(252) * 100,
                        'Sharpe Ratio': (ticker_returns.mean() / ticker_returns.std()) * np.sqrt(252) if ticker_returns.std() > 0 else 0,
                        'Max Drawdown (%)': ((1 + ticker_returns).cumprod().div((1 + ticker_returns).cumprod().cummax()) - 1).min() * 100,
                        'Best Day (%)': ticker_returns.max() * 100,
                        'Worst Day (%)': ticker_returns.min() * 100
                    }
                    portfolio_stats.append(stats)

                portfolio_df = pd.DataFrame(portfolio_stats).round(2)
                analysis_data['Portfolio_Statistics'] = portfolio_df

                st.dataframe(portfolio_df, use_container_width=True, hide_index=True)

                col1, col2 = st.columns(2)
                with col1:
                    best_performer = portfolio_df.loc[portfolio_df['Total Return (%)'].idxmax()]
                    st.success(f"🏆 **Best Performer**: {best_performer['Ticker']} ({best_performer['Total Return (%)']:+.2f}%)")
                with col2:
                    worst_performer = portfolio_df.loc[portfolio_df['Total Return (%)'].idxmin()]
                    st.error(f"📉 **Worst Performer**: {worst_performer['Ticker']} ({worst_performer['Total Return (%)']:+.2f}%)")

        if include_earnings_metrics and metrics_data:
            st.divider()
            st.markdown("### 💰 Earnings vs Metrics Preview")
            preview_metrics_ticker = st.selectbox("Select ticker to preview fundamentals:",
                                                    options=list(metrics_data.keys()), key="metrics_preview_select")
            if preview_metrics_ticker:
                st.dataframe(metrics_data[preview_metrics_ticker], use_container_width=True)
                snapshot = metrics_data[preview_metrics_ticker].attrs.get('snapshot', {})
                if snapshot:
                    st.caption("Current snapshot & analyst estimates:")
                    st.write(snapshot)

        if compare_benchmark and benchmark_data:
            st.divider()
            st.markdown(f"### 📈 Benchmark Comparison Preview ({benchmark_ticker})")
            preview_bench_ticker = st.selectbox("Select ticker to preview vs benchmark:",
                                                 options=list(benchmark_data.keys()), key="benchmark_preview_select")
            if preview_bench_ticker:
                bres = benchmark_data[preview_bench_ticker]
                cols = st.columns(5)
                stat_items = list(bres['stats'].items())
                for c, (label, val) in zip(cols, stat_items):
                    c.metric(label, f"{val:.2f}" if pd.notna(val) else "N/A")
                st.dataframe(bres['rel_df'].tail(10), use_container_width=True)

        st.divider()
        st.markdown("### 📋 Data Preview")
        preview_ticker = st.selectbox("Select ticker to preview:", options=list(ticker_data.keys()),
                                       format_func=lambda x: f"{x} ({len(ticker_data[x])} records)")

        if preview_ticker:
            with st.expander(f"📊 {preview_ticker} - Preview", expanded=True):
                preview_df = ticker_data[preview_ticker]
                col1, col2 = st.columns(2)
                with col1:
                    st.write("**First 5 rows:**")
                    st.dataframe(preview_df.head(5), use_container_width=True)
                with col2:
                    st.write("**Last 5 rows:**")
                    st.dataframe(preview_df.tail(5), use_container_width=True)

                if 'Close' in preview_df.columns:
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Start Price", f"${preview_df['Close'].iloc[0]:.2f}")
                    with col2:
                        st.metric("End Price", f"${preview_df['Close'].iloc[-1]:.2f}")
                    with col3:
                        total_ret = ((preview_df['Close'].iloc[-1] / preview_df['Close'].iloc[0] - 1) * 100)
                        st.metric("Total Return", f"{total_ret:+.2f}%")
                    with col4:
                        st.metric("Data Points", len(preview_df))

        st.divider()
        st.markdown("### 📥 Download Excel Report")

        with st.spinner("🔨 Generating Excel workbook with formatting..."):
            metadata = {
                'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'tickers': list(ticker_data.keys()),
                'start_date': start_date,
                'end_date': end_date,
                'frequency': frequency,
                'total_records': total_records
            }

            exporter = ExcelExporter()
            excel_buffer = exporter.create_workbook(
                ticker_data=ticker_data,
                analysis_data=analysis_data if include_analytics else None,
                metadata=metadata,
                include_charts=include_charts and OPENPYXL_AVAILABLE,
                metrics_data=metrics_data if include_earnings_metrics else None,
                dividends_data=dividends_data if include_dividends else None,
                benchmark_data=benchmark_data if compare_benchmark else None,
                seasonality_data=seasonality_data if include_seasonality else None,
                profile_data=profile_data if include_profile else None,
            )

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if len(ticker_data) == 1:
            file_name = f"{list(ticker_data.keys())[0]}_{frequency.lower()}_{start_date}_to_{end_date}.xlsx"
        else:
            file_name = f"stock_data_{len(ticker_data)}tickers_{frequency.lower()}_{timestamp}.xlsx"

        st.download_button(
            label=f"📥 Download Excel Report • {len(ticker_data)} ticker(s) • {total_records:,} records",
            data=excel_buffer, file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )

        n_metrics_sheets = len(metrics_data) if include_earnings_metrics else 0
        n_div_sheets = len(dividends_data) if include_dividends else 0
        n_bench_sheets = len(benchmark_data) if compare_benchmark else 0
        n_season_sheets = len(seasonality_data) if include_seasonality else 0
        n_corr_sheets = 1 if len(ticker_data) > 1 else 0
        total_sheets = (2 + len(ticker_data) + len(analysis_data) + (1 if len(ticker_data) > 1 else 0)
                        + n_metrics_sheets + n_div_sheets + n_bench_sheets + n_season_sheets + n_corr_sheets)

        st.success("✅ Excel file ready for download!")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📑 Total Sheets", total_sheets)
        with col2:
            st.metric("💾 Estimated Size", f"~{(total_records * 50) / 1024:.1f} KB")
        with col3:
            st.metric("📊 Charts", "Yes" if include_charts and OPENPYXL_AVAILABLE else "No")

        with st.expander("📑 Detailed Sheet Breakdown"):
            sheets_list = ["1. **Contents** - Hyperlinked index / cover page",
                            "2. **Summary** - Export metadata, performance overview, company snapshot & key takeaways"]

            for i, ticker in enumerate(ticker_data.keys(), 3):
                sheets_list.append(f"{i}. **{ticker}_Prices** - Historical price data with SMA/RSI, {len(ticker_data[ticker])} records")

            sheet_num = len(ticker_data) + 3
            for sheet_name in analysis_data.keys():
                sheets_list.append(f"{sheet_num}. **{sheet_name}** - Analytics and calculations")
                sheet_num += 1

            if len(ticker_data) > 1:
                sheets_list.append(f"{sheet_num}. **Comparison** - Multi-ticker comparison + Growth of $100 chart")
                sheet_num += 1
                sheets_list.append(f"{sheet_num}. **Correlation** - Return correlation heatmap")
                sheet_num += 1

            if include_earnings_metrics:
                for ticker in metrics_data.keys():
                    sheets_list.append(f"{sheet_num}. **{ticker}_Earnings_Metrics** - Earnings vs fundamentals + analyst estimates")
                    sheet_num += 1

            if include_dividends:
                for ticker in dividends_data.keys():
                    sheets_list.append(f"{sheet_num}. **{ticker}_Dividends** - Dividend & split history")
                    sheet_num += 1

            if compare_benchmark:
                for ticker in benchmark_data.keys():
                    sheets_list.append(f"{sheet_num}. **{ticker}_vs_{benchmark_ticker}** - Relative performance, beta, alpha")
                    sheet_num += 1

            if include_seasonality:
                for ticker in seasonality_data.keys():
                    sheets_list.append(f"{sheet_num}. **{ticker}_Seasonality** - Avg return by month & day of week")
                    sheet_num += 1

            for sheet_desc in sheets_list:
                st.markdown(sheet_desc)


if __name__ == "__main__":
    st.set_page_config(page_title="Enhanced Excel Export", layout="wide", page_icon="📥")
    render_excel_export()
